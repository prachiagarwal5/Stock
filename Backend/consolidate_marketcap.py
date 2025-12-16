import pandas as pd
import glob
import os
import json
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re

class MarketCapConsolidator:
    def __init__(self, data_folder, output_file='Finished_Product.xlsx', config_file='corporate_actions.json', file_type='mcap'):
        self.data_folder = data_folder
        self.output_file = os.path.join(data_folder, output_file)
        self.config_file = os.path.join(data_folder, config_file)
        self.df_consolidated = None
        self.dates_list = []
        self.file_type = file_type  # 'mcap' or 'pr'
        self.corporate_actions = self._load_corporate_actions()
        self._detect_columns()
    
    def _detect_columns(self):
        """Detect which columns to use based on file type"""
        if self.file_type == 'pr':
            # PR files use: SECURITY, NET_TRDQTY
            self.symbol_col = 'SECURITY'
            self.value_col = 'NET_TRDQTY'
            self.name_col = 'SECURITY'  # Use SECURITY for company name in PR files
        else:
            # MCAP files use: Symbol, Market Cap(Rs.)
            self.symbol_col = 'Symbol'
            self.value_col = 'Market Cap(Rs.)'
            self.name_col = 'Security Name'
    
    def _load_corporate_actions(self):
        """Load corporate actions configuration"""
        if os.path.exists(self.config_file):
            with open(self.config_file, 'r') as f:
                return json.load(f)
        return {
            "splits": [],
            "name_changes": [],
            "delistings": []
        }
    
    def _extract_date_from_filename(self, filename):
        """Extract date from filename like mcap10112025.csv or pr10112025.csv"""
        # Try mcap pattern first
        match = re.search(r'mcap(\d{8})', filename)
        if match:
            date_str = match.group(1)
            # Format: DDMMYYYY
            day = date_str[0:2]
            month = date_str[2:4]
            year = date_str[4:8]
            return f"{day}-{month}-{year}"
        
        # Try pr pattern
        match = re.search(r'pr(\d{8})', filename)
        if match:
            date_str = match.group(1)
            # Format: DDMMYYYY
            day = date_str[0:2]
            month = date_str[2:4]
            year = date_str[4:8]
            return f"{day}-{month}-{year}"
        
        return None
    
    def _parse_date_string(self, date_str):
        """Parse date string like '10-11-2025' to datetime"""
        try:
            return datetime.strptime(date_str, "%d-%m-%Y")
        except:
            return None
    
    def load_and_consolidate_data(self):
        """Load all CSV files and consolidate data"""
        # Choose file pattern based on type
        if self.file_type == 'pr':
            pattern = os.path.join(self.data_folder, 'pr*.csv')
        else:
            pattern = os.path.join(self.data_folder, 'mcap*.csv')
        
        csv_files = sorted(glob.glob(pattern))
        
        if not csv_files:
            print(f"No {self.file_type.upper()} CSV files found in {self.data_folder}")
            return 0, 0  # Return 0, 0 instead of False
        
        print(f"Found {len(csv_files)} {self.file_type.upper()} CSV files")
        
        # Dictionary to store data: {symbol: {date: value}}
        consolidated_data = {}
        company_names = {}  # Store company names
        self.dates_list = []
        
        for csv_file in csv_files:
            date_str = self._extract_date_from_filename(os.path.basename(csv_file))
            if not date_str:
                continue
            
            self.dates_list.append((date_str, self._parse_date_string(date_str)))
            print(f"Processing: {os.path.basename(csv_file)} ({date_str})")
            
            # Read CSV
            try:
                df = pd.read_csv(csv_file)
                # Strip whitespace from column names
                df.columns = df.columns.str.strip()
            except Exception as e:
                print(f"Error reading {csv_file}: {e}")
                continue
            
            # Check if required columns exist
            if self.symbol_col not in df.columns:
                print(f"Warning: Column '{self.symbol_col}' not found. Available columns: {list(df.columns)}")
                continue
            
            if self.value_col not in df.columns:
                print(f"Warning: Column '{self.value_col}' not found. Available columns: {list(df.columns)}")
                continue
            
            # Extract Symbol and Value (Market Cap or NET_TRDVAL)
            for idx, row in df.iterrows():
                symbol = str(row.get(self.symbol_col, '')).strip()
                value = row.get(self.value_col, '')
                company_name = str(row.get(self.name_col, symbol)).strip()
                
                if symbol and value:
                    if symbol not in consolidated_data:
                        consolidated_data[symbol] = {}
                    
                    # Convert value to float
                    try:
                        value_float = float(value)
                        consolidated_data[symbol][date_str] = value_float
                    except:
                        consolidated_data[symbol][date_str] = None
                    
                    company_names[symbol] = company_name
        
        # Sort dates
        self.dates_list.sort(key=lambda x: x[1])
        sorted_dates = [d[0] for d in self.dates_list]
        
        # Create consolidated dataframe
        data_for_df = []
        for symbol in sorted(consolidated_data.keys()):
            row = {
                'Symbol': symbol,
                'Company Name': company_names.get(symbol, symbol)
            }
            
            # Add values for each date
            for date_str in sorted_dates:
                row[date_str] = consolidated_data[symbol].get(date_str, None)
            
            data_for_df.append(row)
        
        self.df_consolidated = pd.DataFrame(data_for_df)
        companies_count = len(self.df_consolidated)
        dates_count = len(sorted_dates)
        print(f"\nConsolidated data: {companies_count} companies across {dates_count} dates")
        return companies_count, dates_count
    
    def apply_corporate_actions(self):
        """Apply corporate actions to blank out cells before split dates"""
        if self.df_consolidated is None:
            return
        
        print("\nApplying corporate actions...")
        
        # Handle stock splits
        for split in self.corporate_actions.get('splits', []):
            old_symbol = split.get('old_symbol', '').strip()
            new_symbols = split.get('new_symbols', [])
            split_date = split.get('split_date', '')
            
            if old_symbol in self.df_consolidated['Symbol'].values:
                print(f"Processing split: {old_symbol} -> {new_symbols} on {split_date}")
                
                # Find columns before split date
                date_columns = [col for col in self.df_consolidated.columns 
                              if col not in ['Symbol', 'Company Name']]
                
                split_datetime = self._parse_date_string(split_date)
                if split_datetime:
                    for date_col in date_columns:
                        col_datetime = self._parse_date_string(date_col)
                        if col_datetime and col_datetime < split_datetime:
                            # Blank out old symbol before split
                            self.df_consolidated.loc[
                                self.df_consolidated['Symbol'] == old_symbol, 
                                date_col
                            ] = None
        
        # Handle name changes
        for change in self.corporate_actions.get('name_changes', []):
            old_symbol = change.get('old_symbol', '').strip()
            new_symbol = change.get('new_symbol', '').strip()
            change_date = change.get('change_date', '')
            
            print(f"Processing name change: {old_symbol} -> {new_symbol} on {change_date}")
            
            # Rename symbol from change_date onwards
            if old_symbol in self.df_consolidated['Symbol'].values:
                date_columns = [col for col in self.df_consolidated.columns 
                              if col not in ['Symbol', 'Company Name']]
                
                change_datetime = self._parse_date_string(change_date)
                if change_datetime:
                    for date_col in date_columns:
                        col_datetime = self._parse_date_string(date_col)
                        if col_datetime and col_datetime < change_datetime:
                            self.df_consolidated.loc[
                                self.df_consolidated['Symbol'] == old_symbol,
                                date_col
                            ] = None
    
    def format_excel_output(self, output_file=None):
        """Format and save to Excel with styling"""
        if output_file:
            self.output_file = output_file
        
        print(f"\nCreating Excel file: {self.output_file}")
        
        # Replace NaN with None for cleaner output
        df = self.df_consolidated.copy()
        df = df.where(pd.notna(df), None)
        
        # Save to temporary CSV first
        temp_csv = self.output_file.replace('.xlsx', '_temp.csv')
        df.to_csv(temp_csv, index=False)
        
        # Create Excel workbook with formatting
        wb = Workbook()
        ws = wb.active
        # Title based on file type
        if self.file_type == 'pr':
            ws.title = "Net Traded Qty"
        else:
            ws.title = "Market Cap Data"
        
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        headers = df.columns.tolist()
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = str(header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Write data
        number_format = '#,##0.00'
        data_alignment = Alignment(horizontal="right", vertical="center")
        
        for row_num, row in enumerate(df.values, 2):
            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num)
                
                if col_num <= 2:  # Symbol and Company Name columns
                    # Convert to string
                    cell.value = str(value) if value is not None else ""
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:  # Date columns with values
                    if value is not None and value != "" and not (isinstance(value, float) and pd.isna(value)):
                        try:
                            # Try to convert to float for numeric values
                            numeric_value = float(value)
                            cell.value = numeric_value
                            cell.number_format = number_format
                            cell.alignment = data_alignment
                        except (ValueError, TypeError):
                            cell.value = value
                            cell.alignment = data_alignment
                    else:
                        cell.value = None
                
                cell.border = border
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15  # Symbol
        ws.column_dimensions['B'].width = 30  # Company Name
        
        for col_num in range(3, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 18
        
        # Freeze panes (first row and first two columns)
        ws.freeze_panes = "C2"
        
        wb.save(self.output_file)
        print(f"✓ Excel file created: {self.output_file}")
        
        # Cleanup temp file
        if os.path.exists(temp_csv):
            os.remove(temp_csv)
    
    def create_corporate_actions_template(self):
        """Create a template for corporate actions if it doesn't exist"""
        if not os.path.exists(self.config_file):
            template = {
                "splits": [
                    {
                        "old_symbol": "TATAMOTOR",
                        "new_symbols": ["TMPV", "TMCV"],
                        "split_date": "DD-MM-YYYY",
                        "description": "Stock split/demerger - blank old symbol before this date"
                    }
                ],
                "name_changes": [
                    {
                        "old_symbol": "OLDNAME",
                        "new_symbol": "NEWNAME",
                        "change_date": "DD-MM-YYYY",
                        "description": "Company name change - blank old symbol before this date"
                    }
                ],
                "delistings": [
                    {
                        "symbol": "DELISTED",
                        "delisting_date": "DD-MM-YYYY",
                        "description": "Company delisted - blank from this date onwards"
                    }
                ]
            }
            
            with open(self.config_file, 'w') as f:
                json.dump(template, f, indent=2)
            print(f"✓ Created corporate actions template: {self.config_file}")
    
    def run(self):
        """Run the complete consolidation process"""
        file_type_name = "Net Traded Value" if self.file_type == 'pr' else "Market Cap"
        print("=" * 60)
        print(f"{file_type_name} Consolidation Tool")
        print("=" * 60)
        
        self.create_corporate_actions_template()
        
        companies_count, dates_count = self.load_and_consolidate_data()
        if companies_count > 0:
            self.apply_corporate_actions()
            self.format_excel_output()
            print("\n✓ Consolidation complete!")
            return True
        else:
            print("✗ Consolidation failed!")
            return False

def main():
    # Set the data folder
    data_folder = '/Users/vinayak/Desktop/Proj01/nosubject'
    
    consolidator = MarketCapConsolidator(data_folder)
    consolidator.run()

if __name__ == "__main__":
    main()
