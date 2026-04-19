from flask import Flask, request, jsonify, send_file, Response
from flask_cors import CORS
import os
import pandas as pd
import glob
import json
import csv
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
from werkzeug.utils import secure_filename
import tempfile
import shutil
from pathlib import Path
import requests
from io import BytesIO, StringIO
import zipfile
import time
from dateutil import parser as date_parser
import numpy as np
from pymongo import MongoClient, UpdateOne
from bson.binary import Binary
from bson import ObjectId
import dotenv
import base64
import math
from consolidate_marketcap import MarketCapConsolidator
from nse_symbol_metrics import SymbolMetricsFetcher
from urllib.parse import quote_plus
from concurrent.futures import ThreadPoolExecutor, as_completed
from memory_optimized_export import MemoryOptimizedExporter, ChunkedDataProcessor, get_memory_usage_mb
import gc
from auth_routes import auth_bp, init_auth

app = Flask(__name__)

# Process start marker used by keep-alive endpoint for lightweight diagnostics
PROCESS_START_TIME = datetime.now()

# Enhanced CORS configuration for production deployment
CORS(app, 
     resources={r"/*": {
         "origins": "*",
         "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"],
         "allow_headers": ["Content-Type", "Authorization", "Accept", "X-Requested-With"],
         "expose_headers": ["Content-Disposition", "X-Export-Log"],
         "supports_credentials": False,
         "max_age": 3600
     }})

@app.after_request
def add_cors_headers(response):
    # Ensure CORS headers are present (fallback for production)
    if 'Access-Control-Allow-Origin' not in response.headers:
        response.headers['Access-Control-Allow-Origin'] = '*'
    
    if 'Access-Control-Allow-Methods' not in response.headers:
        response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS'
    
    if 'Access-Control-Allow-Headers' not in response.headers:
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization, Accept, X-Requested-With'
    
    # Add expose headers
    expose = response.headers.get('Access-Control-Expose-Headers', '')
    needed = ['Content-Disposition', 'X-Export-Log']
    if expose:
        existing = [h.strip() for h in expose.split(',') if h.strip()]
    else:
        existing = []
    for h in needed:
        if h not in existing:
            existing.append(h)
    response.headers['Access-Control-Expose-Headers'] = ', '.join(existing)
    
    return response

@app.before_request
def handle_preflight():
    """Handle OPTIONS preflight requests"""
    if request.method == 'OPTIONS':
        response = app.make_default_options_response()
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization, Accept, X-Requested-With'
        response.headers['Access-Control-Max-Age'] = '3600'
        return response

# Register auth blueprint
app.register_blueprint(auth_bp)

@app.route("/")
def home():
    return "Backend is running 🚀"

# Load environment variables
dotenv.load_dotenv()

# Keep-alive security settings (for GitHub Actions scheduled pings)
KEEPALIVE_TOKEN = os.getenv('KEEPALIVE_TOKEN', '').strip()
KEEPALIVE_ALLOW_UNAUTH = os.getenv('KEEPALIVE_ALLOW_UNAUTH', 'true').strip().lower() == 'true'


def _is_keepalive_authorized(req):
    """Validate keep-alive token if token auth is enabled."""
    # If unauthenticated access is explicitly allowed, accept request.
    if KEEPALIVE_ALLOW_UNAUTH:
        return True

    if not KEEPALIVE_TOKEN:
        # No token configured while unauthenticated access is disabled.
        return False

    provided = (
        req.headers.get('X-Keepalive-Token')
        or req.args.get('token')
        or req.headers.get('Authorization', '').replace('Bearer ', '', 1)
    )
    return bool(provided) and provided.strip() == KEEPALIVE_TOKEN

# MongoDB connection
try:
    # Hardcoded MongoDB connection string (as per user request)
    mongo_uri = "mongodb+srv://prachiagrawal509:BSzCRUTG8F7voUBv@cluster0.kfbej.mongodb.net/Stocks?retryWrites=true&w=majority"
    
    print(f"🔄 Connecting to MongoDB...")
    mongo_client = MongoClient(mongo_uri, serverSelectionTimeoutMS=10000)
    # Test the connection
    mongo_client.admin.command('ping')
    db = mongo_client['Stocks']
    excel_results_collection = db['excel_results']
    bhavcache_collection = db['bhavcache']
    symbol_daily_collection = db['symbol_daily']  # per-symbol, per-date values (mcap/pr)
    symbol_aggregates_collection = db['symbol_aggregates']  # per-symbol averages
    symbol_metrics_collection = db['symbol_metrics']  # Symbol dashboard metrics
    symbol_metrics_daily_collection = db['symbol_metrics_daily']  # Daily impact_cost and free_float_mcap
    nifty_indices_collection = db['nifty_indices']  # Nifty index constituent mappings
    
    print(f"🔄 Creating indexes...")
    # speed-critical indexes
    symbol_daily_collection.create_index(
        [('symbol', 1), ('type', 1), ('date', 1)], name='symbol_type_date', unique=True
    )
    symbol_daily_collection.create_index([('type', 1), ('date', 1)], name='type_date')
    symbol_aggregates_collection.create_index(
        [('symbol', 1), ('type', 1), ('date_range.start', 1), ('date_range.end', 1)],
        name='symbol_type_range', unique=False
    )
    bhavcache_collection.create_index([('type', 1), ('date', 1)], name='type_date_cache')
    
    # New schema: one document per symbol with date-wise nested data
    # Drop old collection to migrate to new schema if duplicates exist
    try:
        symbol_metrics_daily_collection.create_index([('symbol', 1)], name='symbol_idx', unique=True)
    except Exception as idx_err:
        if 'duplicate key' in str(idx_err).lower() or 'E11000' in str(idx_err):
            print(f"⚠️ Found old schema data with duplicates. Migrating to new schema...")
            print(f"⚠️ Dropping old symbol_metrics_daily collection data...")
            symbol_metrics_daily_collection.drop()
            print(f"✓ Collection dropped. Creating fresh indexes...")
            symbol_metrics_daily_collection.create_index([('symbol', 1)], name='symbol_idx', unique=True)
        else:
            raise
    
    symbol_metrics_daily_collection.create_index([('daily_data.date', 1)], name='daily_date_idx')
    nifty_indices_collection.create_index([('symbol', 1)], name='symbol_idx', unique=True)
    print("✅ MongoDB connected successfully")
    # Initialise auth (users collection + indexes) — does NOT touch any existing collection
    init_auth(db)
except Exception as e:
    print(f"⚠️ MongoDB connection failed: {e}")
    import traceback
    traceback.print_exc()
    db = None
    excel_results_collection = None
    bhavcache_collection = None
    symbol_daily_collection = None
    symbol_aggregates_collection = None
    symbol_metrics_collection = None
    symbol_metrics_daily_collection = None
    nifty_indices_collection = None


# Custom JSON encoder to handle NaN and Inf values
class NumpyEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, np.floating):
            if np.isnan(obj) or np.isinf(obj):
                return None
            return float(obj)
        elif isinstance(obj, np.integer):
            return int(obj)
        elif isinstance(obj, np.ndarray):
            return obj.tolist()
        return super().default(obj)

def convert_nan_to_none(obj):
    """Recursively convert NaN, inf, and other non-serializable values to None"""
    if isinstance(obj, float):
        if np.isnan(obj) or np.isinf(obj):
            return None
        return obj
    elif isinstance(obj, (list, tuple)):
        return [convert_nan_to_none(item) for item in obj]
    elif isinstance(obj, dict):
        return {key: convert_nan_to_none(value) for key, value in obj.items()}
    elif isinstance(obj, pd.Series):
        return obj.where(pd.notna(obj), None).to_list()
    elif isinstance(obj, pd.DataFrame):
        return obj.where(pd.notna(obj), None).values.tolist()
    return obj


def is_summary_symbol(symbol):
    """Detect summary rows such as Total/Listed (with spaces or punctuation)."""
    if symbol is None:
        return False
    text = str(symbol).strip().upper()
    if not text:
        return False
    normalized = re.sub(r'[^A-Z0-9]', '', text)
    summary_tokens = {'TOTAL', 'LISTED', 'TOTALLISTED', 'LISTEDTOTAL'}
    if normalized in summary_tokens:
        return True
    if text.startswith('TOTAL') or text.startswith('LISTED'):
        return True
    return False


# ===== Index utilities =====
def _make_session(user_agent=None):
    sess = requests.Session()
    if user_agent:
        sess.headers.update({'User-Agent': user_agent})
    return sess


def _prime_cookies(session, headers):
    try:
        session.get('https://www.nseindia.com', headers=headers, timeout=10)
    except Exception as exc:  # best-effort warmup
        print(f"⚠️ NSE cookie warmup failed (indices): {exc}")


def fetch_index_constituents(index_name, session, headers):
    url = f"https://www.nseindia.com/api/equity-stock?index={quote_plus(index_name)}"
    resp = session.get(url, headers=headers, timeout=20)
    if resp.status_code != 200:
        raise ValueError(f"Index fetch failed {resp.status_code} for {index_name}")
    try:
        data = resp.json() if resp.content else {}
    except Exception:
        raise ValueError(f"Invalid JSON for index {index_name}")
    rows = data.get('data') or []
    def clean_num(v):
        if v is None or v == "" or v == "-": return None
        if isinstance(v, (int, float)): return float(v)
        try:
            return float(str(v).replace(',', '').strip())
        except:
            return None

    symbol_data_map = {}
    for row in rows:
        sym = row.get('symbol') or row.get('symbolName') or row.get('securitySymbol')
        if sym:
            sym = str(sym).strip().upper()
            # MC is Total Market Cap (expressed in Cr or full? Usually Cr in this API)
            # FF is Free Float Market Cap
            symbol_data_map[sym] = {
                'mc': clean_num(row.get('marketCap')),
                'ff': clean_num(row.get('ffmc'))
            }
    return symbol_data_map


def build_symbol_index_map(index_list):
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
        'Accept': 'application/json,text/plain,*/*',
        'Connection': 'keep-alive',
        'Referer': 'https://www.nseindia.com/market-data/live-market-indices'
    }
    sess = _make_session()
    _prime_cookies(sess, headers)
    mapping = {}
    live_data_mapping = {}
    errors = []
    for idx in index_list:
        try:
            index_data = fetch_index_constituents(idx, sess, headers)
            for sym, data in index_data.items():
                if sym not in mapping:
                    mapping[sym] = [idx]
                    live_data_mapping[sym] = data
                else:
                    if idx not in mapping[sym]:
                        mapping[sym].append(idx)
        except Exception as exc:
            errors.append({'index': idx, 'error': str(exc)})
    return mapping, live_data_mapping, errors


def primary_index_map_from_db(symbols):
    """Return latest primary_index per symbol from nifty_indices_collection."""
    if nifty_indices_collection is None or not symbols:
        return {}
    mapping = {}
    try:
        # Query exactly what we need: current constituents only
        cursor = nifty_indices_collection.find(
            {'symbol': {'$in': symbols}},
            {'symbol': 1, 'primary_index': 1}
        )

        for doc in cursor:
            sym = doc.get('symbol')
            if not sym:
                continue
            idx = doc.get('primary_index')
            if idx:
                mapping[sym] = idx
    except Exception as exc:
        print(f"⚠️ Failed to build primary index map from nifty_indices: {exc}")
    return mapping


def _safe_float(val):
    try:
        if val in (None, '', 'NA', 'NaN'):
            return None
        return float(val)
    except Exception:
        return None


def upsert_symbol_daily(symbol, company_name, date_iso, data_type, value, source='consolidation', extra=None):
    if symbol_daily_collection is None or not date_iso:
        return
    try:
        payload = {
            'symbol': symbol,
            'company_name': company_name,
            'date': date_iso,
            'type': data_type,
            'value': _safe_float(value),
            'source': source,
            'updated_at': datetime.now().isoformat()
        }
        if extra:
            payload.update(extra)
        symbol_daily_collection.update_one(
            {'symbol': symbol, 'date': date_iso, 'type': data_type},
            {'$set': payload},
            upsert=True
        )
    except Exception as exc:
        print(f"⚠️ Failed to upsert symbol_daily for {symbol} {date_iso} {data_type}: {exc}")


def upsert_symbol_aggregate(symbol, company_name, data_type, days_with_data, average_value, date_range=None, source='consolidation'):
    if symbol_aggregates_collection is None:
        return
    try:
        payload = {
            'symbol': symbol,
            'company_name': company_name,
            'type': data_type,
            'days_with_data': int(days_with_data or 0),
            'average': _safe_float(average_value),
            'date_range': date_range,
            'source': source,
            'updated_at': datetime.now().isoformat()
        }
        # Use only symbol + type as the key to avoid duplicates across date ranges
        # Each symbol will have exactly one MCAP record and one PR record
        symbol_aggregates_collection.update_one(
            {'symbol': symbol, 'type': data_type},
            {'$set': payload},
            upsert=True
        )
    except Exception as exc:
        print(f"⚠️ Failed to upsert symbol_aggregate for {symbol} {data_type}: {exc}")


def bulk_upsert_symbol_metrics(rows, source='nse_symbol_metrics'):
    """
    ULTRA-FAST bulk upsert of symbol metrics.
    Consolidates thousands of database operations into 2-3 bulk writes.
    """
    if symbol_metrics_collection is None:
        return
    
    from pymongo import UpdateOne
    ops_main = []
    ops_daily_meta = []
    ops_daily_pull = []
    ops_daily_push = []
    
    now_iso = datetime.now().isoformat()
    default_as_on = datetime.now().strftime('%Y-%m-%d')
    
    try:
        for row in rows:
            symbol = str(row.get('symbol') or '').strip()
            if not symbol: continue
            
            payload = dict(row)
            as_on = payload.get('as_on') or payload.get('date') or default_as_on
            if hasattr(as_on, 'strftime'): as_on = as_on.strftime('%Y-%m-%d')
            
            payload['as_on'] = as_on
            payload['source'] = source
            payload['updated_at'] = now_iso
            
            # 1. Main collection ops
            ops_main.append(UpdateOne(
                {'symbol': symbol, 'as_on': as_on},
                {'$set': payload},
                upsert=True
            ))
            
            # 2. Daily collection (Metadata & Structure)
            if symbol_metrics_daily_collection is not None:
                cname = row.get('companyName') or row.get('company_name') or ''
                daily_entry = {
                    'date': as_on,
                    'impact_cost': _safe_float(row.get('impact_cost')),
                    'free_float_mcap': _safe_float(row.get('free_float_mcap')),
                    'total_market_cap': _safe_float(row.get('total_market_cap')),
                    'total_traded_value': _safe_float(row.get('total_traded_value')),
                    'source': source,
                    'updated_at': now_iso
                }
                
                ops_daily_meta.append(UpdateOne(
                    {'symbol': symbol},
                    {
                        '$set': {'company_name': cname, 'last_updated': now_iso},
                        '$setOnInsert': {'symbol': symbol, 'created_at': now_iso}
                    },
                    upsert=True
                ))
                
                ops_daily_pull.append(UpdateOne(
                    {'symbol': symbol},
                    {'$pull': {'daily_data': {'date': as_on}}}
                ))
                
                ops_daily_push.append(UpdateOne(
                    {'symbol': symbol},
                    {'$push': {'daily_data': daily_entry}}
                ))

        # Execute Bulk Operations
        if ops_main:
            symbol_metrics_collection.bulk_write(ops_main, ordered=False)
        
        if symbol_metrics_daily_collection is not None:
            if ops_daily_meta:
                symbol_metrics_daily_collection.bulk_write(ops_daily_meta, ordered=False)
            if ops_daily_pull:
                symbol_metrics_daily_collection.bulk_write(ops_daily_pull, ordered=False)
            if ops_daily_push:
                symbol_metrics_daily_collection.bulk_write(ops_daily_push, ordered=False)
                
        print(f"✅ Bulk upserted {len(rows)} symbol metrics successfully ({source})")
    except Exception as exc:
        print(f"⚠️ Bulk upsert failed: {exc}")

def upsert_symbol_metrics(row, source='nse_symbol_metrics'):
    """Legacy wrapper for single row upsert"""
    bulk_upsert_symbol_metrics([row], source=source)

METRIC_FIELDS_FROM_NSE = [
    'impact_cost', 'free_float_mcap', 'total_market_cap', 
    'total_traded_value', 'companyName', 'listingDate', 'basicIndustry'
]

def enrich_rows_from_metrics_db(rows):
    """
    For any row missing impact_cost, free_float_mcap, total_market_cap, total_traded_value, companyName, listingDate
    fall back to the most recent stored value in symbol_metrics_collection.
    This handles stocks where the current NSE API call returns None (e.g. BE-series, illiquid stocks).
    """
    if symbol_metrics_collection is None or not rows:
        return rows

    # Find rows with at least one missing metric field
    needs_fallback = [
        r for r in rows
        if any(r.get(f) is None for f in METRIC_FIELDS_FROM_NSE)
    ]
    if not needs_fallback:
        return rows

    syms = list({r['symbol'] for r in needs_fallback if r.get('symbol')})
    if not syms:
        return rows

    # Fetch latest stored doc per symbol (sort by as_on desc)
    try:
        pipeline = [
            {'$match': {'symbol': {'$in': syms}}},
            {'$sort': {'as_on': -1, 'updated_at': -1}},
            {'$group': {
                '_id': '$symbol',
                'impact_cost':        {'$first': '$impact_cost'},
                'free_float_mcap':    {'$first': '$free_float_mcap'},
                'total_market_cap':   {'$first': '$total_market_cap'},
                'total_traded_value': {'$first': '$total_traded_value'},
                'companyName':        {'$first': '$companyName'},
                'listingDate':        {'$first': '$listingDate'},
                'basicIndustry':      {'$first': '$basicIndustry'},
            }}
        ]
        fallback_map = {}
        for doc in symbol_metrics_collection.aggregate(pipeline):
            sym = doc.pop('_id')
            fallback_map[sym] = doc

        filled = 0
        for row in needs_fallback:
            sym = row.get('symbol')
            fb = fallback_map.get(sym)
            if not fb:
                continue
            for field in METRIC_FIELDS_FROM_NSE:
                if row.get(field) is None and fb.get(field) is not None:
                    row[field] = fb[field]
                    filled += 1

        if filled:
            print(f"[enrich_metrics_db] ✓ Filled {filled} null metric fields for {len(fallback_map)} symbols from history")
    except Exception as exc:
        print(f"[enrich_metrics_db] ⚠️ Fallback query failed: {exc}")

    return rows

def persist_consolidated_results(consolidator, data_type, source='consolidation', skip_daily=False):
    """Store per-symbol per-date values and averages into MongoDB using bulk operations with memory optimization."""
    if consolidator is None or consolidator.df_consolidated is None:
        return
    try:
        from pymongo import UpdateOne
        
        date_cols = [d[0] for d in consolidator.dates_list]
        date_range = None
        if date_cols:
            try:
                start_iso = datetime.strptime(date_cols[0], '%d-%m-%Y').strftime('%Y-%m-%d')
                end_iso = datetime.strptime(date_cols[-1], '%d-%m-%Y').strftime('%Y-%m-%d')
                date_range = {'start': start_iso, 'end': end_iso}
            except Exception:
                date_range = None

        # CHUNKED PROCESSING to avoid OOM
        # OPTIMIZATION: Increased batch size for faster persistence
        SYMBOL_BATCH_SIZE = 1000 
        total_rows = len(consolidator.df_consolidated)
        
        print(f"[persist] Starting chunked persistence for {total_rows} symbols ({data_type})...")
        
        for i in range(0, total_rows, SYMBOL_BATCH_SIZE):
            batch_df = consolidator.df_consolidated.iloc[i : i + SYMBOL_BATCH_SIZE]
            aggregate_ops = []
            daily_ops = []
            
            for _, row in batch_df.iterrows():
                symbol = str(row.get('Symbol') or '').strip()
                company_name = str(row.get('Company Name') or '').strip()
                if not symbol:
                    continue

                # Prepare aggregate upsert
                avg_val = row.get(consolidator.avg_col)
                days_val = row.get(consolidator.days_col)
                # Use unified snake_case keys for internal logic
                non_zero_val = row.get(consolidator.non_zero_days_col) if hasattr(consolidator, 'non_zero_days_col') else row.get('non_zero_days', 0)
                total_possible = row.get('total_possible_days', 0)
                
                if symbol_aggregates_collection is not None:
                    payload = {
                        'symbol': symbol,
                        'company_name': company_name,
                        'type': data_type,
                        'days_with_data': int(days_val or 0),
                        'non_zero_days': int(non_zero_val or 0),
                        'total_possible_days': int(total_possible or 0),
                        'average': _safe_float(avg_val),
                        'date_range': date_range,
                        'source': source,
                        'updated_at': datetime.now().isoformat()
                    }
                    if hasattr(consolidator, 'avg_ff_col'):
                        payload['avg_free_float'] = _safe_float(row.get(consolidator.avg_ff_col))
                    aggregate_ops.append(
                        UpdateOne(
                            {'symbol': symbol, 'type': data_type},
                            {'$set': payload},
                            upsert=True
                        )
                    )

                # Prepare daily upserts
                if symbol_daily_collection is not None and not skip_daily:
                    for date_str in date_cols:
                        val = row.get(date_str)
                        if val in (None, ''):
                            continue
                        try:
                            date_iso = datetime.strptime(date_str, '%d-%m-%Y').strftime('%Y-%m-%d')
                        except Exception:
                            continue
                        
                        payload = {
                            'symbol': symbol,
                            'company_name': company_name,
                            'date': date_iso,
                            'type': data_type,
                            'value': _safe_float(val),
                            'source': source,
                            'updated_at': datetime.now().isoformat()
                        }
                        daily_ops.append(
                            UpdateOne(
                                {'symbol': symbol, 'date': date_iso, 'type': data_type},
                                {'$set': payload},
                                upsert=True
                            )
                        )

            # Execute bulk operations for this batch immediately
            if aggregate_ops and symbol_aggregates_collection is not None:
                symbol_aggregates_collection.bulk_write(aggregate_ops, ordered=False)
            
            if daily_ops and symbol_daily_collection is not None:
                symbol_daily_collection.bulk_write(daily_ops, ordered=False)
            
            # Explicitly clear lists and batch DF to free memory
            del aggregate_ops
            del daily_ops
            del batch_df
            gc.collect()
            
            processed = min(i + SYMBOL_BATCH_SIZE, total_rows)
            print(f"[persist] ✓ Batched persistence: {processed}/{total_rows} symbols processed")

        print(f"[persist] ✓ All results persisted for {data_type}")
            
    except Exception as exc:
        import traceback
        traceback.print_exc()
        print(f"⚠️ Failed to persist consolidated results for {data_type}: {exc}")


def _parse_mcap_date_from_filename(filename):
    match = re.search(r'mcap(\d{2})(\d{2})(\d{4})', filename, re.IGNORECASE)
    if not match:
        return None
    day, month, year = match.groups()
    try:
        return datetime(int(year), int(month), int(day))
    except Exception:
        return None


def collect_symbols_from_files(file_paths):
    symbols = set()
    for path in file_paths:
        if not os.path.exists(path):
            continue
        try:
            df = pd.read_csv(path)
            df.columns = df.columns.str.strip()
            if 'Symbol' in df.columns:
                for sym in df['Symbol'].dropna():
                    sym_clean = str(sym).strip()
                    if sym_clean:
                        symbols.add(sym_clean)
        except Exception as exc:
            print(f"⚠️ Unable to read symbols from {path}: {exc}")
    return sorted(symbols)


def find_mcap_files_in_range(base_dir, start_dt, end_dt):
    pattern = os.path.join(base_dir, 'mcap*.csv')
    files = []
    for path in glob.glob(pattern):
        dt = _parse_mcap_date_from_filename(os.path.basename(path))
        if dt and start_dt <= dt <= end_dt:
            files.append(path)
    return sorted(files)


# ===== Cache helpers =====
def _normalize_iso_date(dt):
    return dt.strftime('%Y-%m-%d')


def get_cached_csv(date_iso, data_type):
    if bhavcache_collection is None:
        return None
    doc = bhavcache_collection.find_one({'date': date_iso, 'type': data_type})
    if not doc:
        return None
    try:
        csv_bytes = doc.get('file_data')
        if not csv_bytes:
            return None
        df = pd.read_csv(BytesIO(csv_bytes))
        return {
            'df': df,
            'records': doc.get('records', len(df)),
            'columns': doc.get('columns', df.columns.tolist()),
            'stored_at': doc.get('stored_at')
        }
    except Exception as e:
        print(f"⚠️ Failed to read cached {data_type} for {date_iso}: {e}")
        return None


def get_cached_csv_bulk(date_iso_list, data_type):
    """Bulk fetch cached CSVs for multiple dates (faster than one-by-one)"""
    if bhavcache_collection is None:
        return {}
    
    try:
        # Single query for all dates
        docs = list(bhavcache_collection.find({
            'date': {'$in': date_iso_list},
            'type': data_type
        }))
        
        results = {}
        for doc in docs:
            date_iso = doc.get('date')
            try:
                csv_bytes = doc.get('file_data')
                if csv_bytes:
                    # Determine columns to load based on data_type
                    symbol_col = 'SECURITY' if data_type == 'pr' else 'Symbol'
                    value_col = 'NET_TRDVAL' if data_type == 'pr' else 'Market Cap(Rs.)'
                    name_col = 'SECURITY' if data_type == 'pr' else 'Security Name'

                    req_cols = [symbol_col, value_col]
                    if symbol_col != name_col: # Only add if different from symbol_col
                        req_cols.append(name_col)
                    
                    # OPTIMIZATION: Load only necessary columns and specify dtypes
                    df = pd.read_csv(
                        BytesIO(csv_bytes),
                        usecols=req_cols,
                        dtype={symbol_col: 'category'} # Category saves memory for repeated symbols
                    )
                    
                    results[date_iso] = {
                        'df': df,
                        'records': doc.get('records', len(df)),
                        'columns': doc.get('columns', df.columns.tolist()),
                        'stored_at': doc.get('stored_at')
                    }
            except Exception as e:
                print(f"⚠️ Failed to read cached {data_type} for {date_iso}: {e}")
        
        return results
    except Exception as e:
        print(f"⚠️ Bulk cache fetch failed for {data_type}: {e}")
        return {}


def get_cached_csv_metadata_bulk(date_iso_list, data_type):
    """LIGHTWEIGHT fetch of cache metadata (NO BLOB) for existence checks."""
    if bhavcache_collection is None:
        return {}
    
    try:
        # SINGLE query for all dates, excluding the heavy file_data blob
        docs = list(bhavcache_collection.find({
            'date': {'$in': date_iso_list},
            'type': data_type
        }, {
            'file_data': 0  # CRITICAL: Exclude the binary blob
        }))
        
        results = {}
        for doc in docs:
            date_iso = doc.get('date')
            results[date_iso] = {
                'records': doc.get('records', 0),
                'stored_at': doc.get('stored_at'),
                'is_cached': True
            }
        
        return results
    except Exception as e:
        print(f"⚠️ Metadata cache fetch failed: {e}")
        return {}


def put_cached_csv(date_iso, data_type, df, source='nse'):
    if bhavcache_collection is None or df is None:
        return None
    try:
        csv_buf = BytesIO()
        df.to_csv(csv_buf, index=False)
        bhavcache_collection.update_one(
            {'date': date_iso, 'type': data_type},
            {
                '$set': {
                    'date': date_iso,
                    'type': data_type,
                    'file_data': csv_buf.getvalue(),
                    'records': len(df),
                    'columns': df.columns.tolist(),
                    'stored_at': datetime.now().isoformat(),
                    'source': source
                }
            },
            upsert=True
        )
        return True
    except Exception as e:
        print(f"⚠️ Failed to cache {data_type} for {date_iso}: {e}")
        return None


def bulk_upsert_symbol_daily_from_df(df, date_iso, data_type, source='nse_download', symbol_name_map=None):
    """Fast upsert of per-symbol values into Mongo, avoids per-row round trips."""
    if symbol_daily_collection is None or df is None or df.empty:
        return
    symbol_col = 'Symbol' if data_type == 'mcap' else 'SECURITY'
    name_col = 'Security Name' if data_type == 'mcap' else 'SECURITY'
    value_col = 'Market Cap(Rs.)' if data_type == 'mcap' else 'NET_TRDVAL'
    
    # Pre-process PR mapping if needed
    pr_lookup = None
    if data_type == 'pr' and symbol_name_map:
        # Build normalized lookup: normalized_name -> ticker_symbol
        def normalize_name(name):
            if not name: return ''
            return ''.join(c.upper() for c in str(name) if c.isalnum() or c.isspace()).strip()
        
        pr_lookup = {}
        for company_name, ticker in symbol_name_map.items():
            norm_name = normalize_name(company_name)
            if norm_name: pr_lookup[norm_name] = ticker

    ops = []
    for _, row in df.iterrows():
        raw_symbol = str(row.get(symbol_col) or '').strip()
        if not raw_symbol or is_summary_symbol(raw_symbol):
            continue
            
        symbol = raw_symbol
        company_name = str(row.get(name_col) or symbol).strip()
        
        # PR Mapping Logic
        if data_type == 'pr' and pr_lookup:
            def normalize_name(name):
                if not name: return ''
                return ''.join(c.upper() for c in str(name) if c.isalnum() or c.isspace()).strip()
            
            norm_pr = normalize_name(raw_symbol) # raw_symbol is the company name in PR
            mapped_ticker = pr_lookup.get(norm_pr)
            if mapped_ticker:
                symbol = mapped_ticker
                # company_name stays as the full name from PR if we want, or we can use it from the map if we had it
            else:
                # If we can't map it, skip it to ensure consistent sorting/averaging with MCAP
                # (since PR without MCAP ticker won't group correctly in the pivot)
                continue
                
        raw_value = row.get(value_col)
        value = pd.to_numeric(raw_value, errors='coerce')
        if pd.isna(value):
            continue
        ops.append(UpdateOne(
            {'symbol': symbol, 'type': data_type, 'date': date_iso},
            {'$set': {
                'symbol': symbol,
                'company_name': company_name,
                'type': data_type,
                'date': date_iso,
                'value': float(value),
                'source': source,
                'updated_at': datetime.now().isoformat()
            }},
            upsert=True
        ))
    if not ops:
        return
    # chunk to keep payload moderate
    chunk_size = 1000
    for i in range(0, len(ops), chunk_size):
        try:
            symbol_daily_collection.bulk_write(ops[i:i + chunk_size], ordered=False)
        except Exception as exc:
            print(f"⚠️ bulk upsert for {data_type} {date_iso} failed: {exc}")


def get_consolidated_metrics_from_db(date_iso_list, data_type, allowed_symbols=None):
    """
    ULTRA-FAST consolidation using MongoDB Aggregation Pipeline on symbol_daily collection.
    Avoids fetching and processing raw CSVs for every request.
    Now supports filtering by allowed_symbols for targeted consolidation (like Dashboard).
    """
    if symbol_daily_collection is None:
        return None
    
    try:
        match_query = {
            'type': data_type,
            'date': {'$in': date_iso_list},
            'symbol': {'$nin': ['Permitted', 'PERMITTED']}
        }
        if allowed_symbols:
            match_query['symbol'] = {'$in': list(allowed_symbols)}
            
        pipeline = [
            {
                '$match': match_query
            },
            {
                '$group': {
                    '_id': '$symbol',
                    'Symbol': {'$first': '$symbol'},
                    'Company Name': {'$first': '$company_name'},
                    'sum_val': {'$sum': '$value'},
                    'count_val': {'$sum': 1},
                    # We also need the daily values for the pivot
                    'daily_data': {
                        '$push': {
                            'date': '$date',
                            'val': '$value'
                        }
                    }
                }
            }
        ]
        
        # Execute aggregation
        cursor = symbol_daily_collection.aggregate(pipeline)
        results = list(cursor)
        
        if not results:
            return None
            
        # Convert to DataFrame
        rows = []
        # Pre-sort date_iso_list to be sure of order
        sorted_iso_range = sorted(date_iso_list)
        
        for res in results:
            daily_entries = res.get('daily_data', [])
            daily_vals = [d.get('val') for d in daily_entries]
            non_zero_days = sum(1 for v in daily_vals if v is not None and v > 0)
            
            # Calculate total_possible_days: 
            # 1. Find the earliest date this symbol appeared in the requested range
            all_dates_found = [d['date'] for d in daily_entries if d.get('date')]
            if all_dates_found:
                first_date = min(all_dates_found)
                # 2. Find its index in the sorted_iso_range
                try:
                    first_idx = sorted_iso_range.index(first_date)
                    total_possible = len(sorted_iso_range) - first_idx
                except ValueError:
                    # Date found but not in our list (unexpected but handleable)
                    total_possible = res['count_val']
            else:
                total_possible = res['count_val']

            row = {
                'Symbol': res['Symbol'],
                'Company Name': res['Company Name'],
                'Average Value': res['sum_val'] / res['count_val'] if res['count_val'] > 0 else 0,
                'Days With Data': res['count_val'],
                'non_zero_days': non_zero_days,
                'total_possible_days': total_possible
            }
            # Flatten daily data for pivoting - fast string manipulation for YYYY-MM-DD -> DD-MM-YYYY
            for daily in res.get('daily_data', []):
                d_iso = daily['date']
                try:
                    # Faster than strptime/strftime: "2026-02-02" -> ["2026", "02", "02"] -> "02-02-2026"
                    parts = d_iso.split('-')
                    if len(parts) == 3:
                        date_key = f"{parts[2]}-{parts[1]}-{parts[0]}"
                        row[date_key] = daily['val']
                except:
                    pass
            rows.append(row)
            
        df = pd.DataFrame(rows)
        return df
    except Exception as e:
        print(f"⚠️ MongoDB aggregation failed for {data_type}: {e}")
        return None


def build_consolidated_from_cache(date_iso_list, data_type, allow_missing=False, log_fn=None, allowed_symbols=None, symbol_name_map=None):
    """Build consolidated dataframe - ULTRA-OPTIMIZED with minimal operations."""
    
    if bhavcache_collection is None:
        error_msg = "Database not connected. Check if MONGODB_URI/mongo_uri is correctly set in environment variables."
        if log_fn: log_fn(f"❌ {error_msg}")
        raise ValueError(error_msg)

    # OPTIMIZATION: Try DB aggregation first (it's much faster)
    if log_fn: log_fn(f"⚡ Attempting high-performance DB aggregation for {data_type.upper()}...")
    df_db = get_consolidated_metrics_from_db(date_iso_list, data_type, allowed_symbols=allowed_symbols)
    if df_db is not None and not df_db.empty:
        if log_fn: log_fn(f"✅ DB aggregation successful ({len(df_db)} records)")
        
        # Identify date columns (DD-MM-YYYY)
        date_cols = [c for c in df_db.columns if re.match(r"\d{2}-\d{2}-\d{4}", c)]
        date_cols = sorted(date_cols, key=lambda d: datetime.strptime(d, '%d-%m-%Y'))
        
        # Determine average column name based on type
        if data_type == 'pr':
            avg_col = 'Average Net Traded Value'
        else:
            avg_col = 'Average Market Cap'
            
        # Rename the generic 'Average Value' column
        df_db = df_db.rename(columns={'Average Value': avg_col})
        
        # Sort by average
        if avg_col in df_db.columns:
            df_db = df_db.sort_values(by=avg_col, ascending=False, na_position='last').reset_index(drop=True)
        
        return df_db, [(d, datetime.strptime(d, '%d-%m-%Y')) for d in date_cols], avg_col

    # Fallback to BULK LOAD from CSV cache
    if log_fn: log_fn(f"ℹ️ DB aggregation failed or no data. Falling back to CSV cache...")
    cached_dict = get_cached_csv_bulk(date_iso_list, data_type)
    
    frames = []
    missing_dates = []
    for date_iso in date_iso_list:
        cached = cached_dict.get(date_iso)
        if not cached or cached.get('df') is None:
            missing_dates.append(date_iso)
            continue
        df = cached['df']
        df.columns = df.columns.str.strip()
        df['_date_iso'] = date_iso
        frames.append(df)

    if missing_dates and not allow_missing:
        error_msg = f"No cached {data_type.upper()} data for dates: {', '.join(missing_dates)}. Please 'Download Range' for these dates first."
        if log_fn: log_fn(f"❌ {error_msg}")
        raise ValueError(error_msg)

    if not frames:
        error_msg = f"No cached {data_type.upper()} data available for the entire requested range ({date_iso_list[0]} to {date_iso_list[-1]})."
        if log_fn: log_fn(f"❌ {error_msg}")
        raise ValueError(error_msg)

    # Concatenate all data at once (fast)
    df_all = pd.concat(frames, ignore_index=True)

    if data_type == 'pr':
        symbol_col = 'SECURITY'
        value_col = 'NET_TRDVAL'
        name_col = 'SECURITY'
        avg_col = 'Average Net Traded Value'
    else:
        symbol_col = 'Symbol'
        value_col = 'Market Cap(Rs.)'
        name_col = 'Security Name'
        avg_col = 'Average Market Cap'

    # OPTIMIZATION: Skip strip() on columns - it's slow. Just convert/clean what matters
    df_all[symbol_col] = df_all[symbol_col].astype(str)
    df_all[name_col] = df_all[name_col].astype(str)
    df_all[value_col] = pd.to_numeric(df_all[value_col], errors='coerce')

    # Remove summary rows (vectorized, no apply())
    symbols_upper = df_all[symbol_col].astype(str).str.upper()
    symbols_normalized = symbols_upper.str.replace(r'[^A-Z0-9]', '', regex=True)
    is_summary = symbols_normalized.isin({'TOTAL', 'LISTED', 'TOTALLISTED', 'LISTEDTOTAL', 'PERMITTED'}) | \
                 symbols_upper.str.startswith(('TOTAL', 'LISTED', 'PERMITTED'))
    df_all = df_all[~is_summary]

    # Filter to allowed symbols if needed
    if allowed_symbols is not None:
        df_all = df_all[df_all[symbol_col].isin(allowed_symbols)]
        if df_all.empty:
            raise ValueError(f"No {data_type.upper()} data for requested symbols")

    # Convert dates once (vectorized)
    df_all['_date_str'] = pd.to_datetime(df_all['_date_iso']).dt.strftime('%d-%m-%Y')
    date_cols = sorted(df_all['_date_str'].unique(), key=lambda d: datetime.strptime(d, '%d-%m-%Y'))

    # ULTRA-FAST RESHAPE: No groupby, just direct unstack with drop_duplicates
    df_all = df_all.drop_duplicates(subset=[symbol_col, '_date_str'], keep='last')
    df_all_pivot = df_all.set_index([symbol_col, '_date_str'])[value_col].unstack(fill_value=None)
    df_all_pivot.reset_index(inplace=True)
    df_all_pivot.rename(columns={symbol_col: 'Symbol'}, inplace=True)

    # Get company names (single pass, handles symbol_col == name_col case)
    if symbol_col == name_col:
        # When symbol and name are the same column, map each to itself
        name_lookup = {sym: sym for sym in df_all[symbol_col].drop_duplicates().values}
    else:
        # Normal case: create mapping from different columns
        unique_df = df_all.drop_duplicates(symbol_col, keep='last')
        name_lookup = unique_df.set_index(symbol_col)[name_col].to_dict()
    df_all_pivot['Company Name'] = df_all_pivot['Symbol'].map(name_lookup)

    # For PR data: Replace company names (in Symbol column) with ticker symbols from MCAP
    if data_type == 'pr' and symbol_name_map:
        # symbol_name_map is CompanyName → TickerSymbol mapping
        # df_all_pivot['Symbol'] currently contains company names (from SECURITY column)
        # Map them to ticker symbols so they match MCAP symbols
        original_companies = df_all_pivot['Symbol'].copy()
        before_count = len(df_all_pivot)
        
        # Create normalized mapping for better matching
        # Normalize: uppercase, remove extra spaces, punctuation
        def normalize_name(name):
            if not name:
                return ''
            return ''.join(c.upper() for c in str(name) if c.isalnum() or c.isspace()).strip()
        
        # Build normalized lookup: normalized_name → ticker_symbol
        normalized_map = {}
        for company_name, ticker in symbol_name_map.items():
            norm_name = normalize_name(company_name)
            if norm_name:
                normalized_map[norm_name] = ticker
        
        # Try exact match first, then normalized match
        def find_ticker(pr_company_name):
            # Try exact match
            if pr_company_name in symbol_name_map:
                return symbol_name_map[pr_company_name]
            # Try normalized match
            norm_pr = normalize_name(pr_company_name)
            return normalized_map.get(norm_pr)
        
        df_all_pivot['Symbol'] = original_companies.apply(find_ticker)
        # Keep original company name
        df_all_pivot['Company Name'] = original_companies
        # Remove rows where mapping failed (no matching MCAP symbol)
        df_all_pivot = df_all_pivot[df_all_pivot['Symbol'].notna()]
        after_count = len(df_all_pivot)
        matched_count = after_count
        unmatched_count = before_count - after_count
        
        if log_fn:
            log_fn(f"✓ PR symbol mapping: {matched_count}/{before_count} matched ({100*matched_count/before_count:.1f}%), {unmatched_count} unmatched")
            if unmatched_count > 0 and unmatched_count <= 10:
                # Show samples of unmatched names
                unmatched_samples = original_companies[~original_companies.isin(df_all_pivot['Company Name'].values)].head(10).tolist()
                log_fn(f"  Unmatched samples: {', '.join(unmatched_samples[:5])}")

    # Get available date columns
    available_cols = [c for c in date_cols if c in df_all_pivot.columns]
    
    # Vectorized metrics (one pass through data)
    numeric_pivot = df_all_pivot[available_cols].apply(pd.to_numeric, errors='coerce') if available_cols else pd.DataFrame()
    df_all_pivot['Days With Data'] = numeric_pivot.notna().sum(axis=1) if not numeric_pivot.empty else 0
    # Non Zero Days: days where value > 0 (actually traded, not just present/non-null)
    df_all_pivot['Non Zero Days'] = (numeric_pivot > 0).sum(axis=1) if not numeric_pivot.empty else 0
    df_all_pivot[avg_col] = numeric_pivot.mean(axis=1) if not numeric_pivot.empty else None

    # Total possible days: count from first data point for each symbol to end of date range
    if not numeric_pivot.empty:
        first_idx = numeric_pivot.notna().values.argmax(axis=1)
        df_all_pivot['total_possible_days'] = len(available_cols) - first_idx
    else:
        df_all_pivot['total_possible_days'] = 0

    # Sort by average
    df_all_pivot = df_all_pivot.sort_values(by=avg_col, ascending=False, na_position='last').reset_index(drop=True)

    # Convert values to floats for safe calculation (no scaling here, raw values stored in Mongo)
    if available_cols:
        for col in available_cols:
            df_all_pivot[col] = pd.to_numeric(df_all_pivot[col], errors='coerce')
        df_all_pivot[avg_col] = pd.to_numeric(df_all_pivot[avg_col], errors='coerce')

    # Final column order
    final_cols = ['Symbol', 'Company Name', 'Days With Data', 'Non Zero Days', 'total_possible_days', avg_col] + available_cols
    df_all_pivot = df_all_pivot[[c for c in final_cols if c in df_all_pivot.columns]]

    # FINAL SAFETY FILTER: Remove 'PERMITTED' symbol if it survived aggregation/CSV parsing
    if 'Symbol' in df_all_pivot.columns:
        df_all_pivot = df_all_pivot[df_all_pivot['Symbol'].astype(str).str.strip().str.upper() != 'PERMITTED'].copy()

    return df_all_pivot, dates_list, avg_col

# Configuration
UPLOAD_FOLDER = os.getenv('UPLOAD_FOLDER', './Backend/uploads/market_cap')
ALLOWED_EXTENSIONS = {'csv'}
MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_FILE_SIZE

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/health', methods=['GET'])
def health():
    """Health check endpoint"""
    return jsonify({
        'status': 'ok',
        'message': 'Market Cap Consolidation Service is running'
    }), 200


@app.route('/api/keepalive', methods=['GET'])
def keepalive():
    """
    Lightweight keep-alive endpoint for external schedulers (e.g., GitHub Actions).

    Suggested use on Render free plan:
    - Ping every 5-10 minutes.
    - Set KEEPALIVE_ALLOW_UNAUTH=false and configure KEEPALIVE_TOKEN for protection.
    """
    if not _is_keepalive_authorized(request):
        return jsonify({'error': 'Unauthorized'}), 401

    uptime_seconds = int((datetime.now() - PROCESS_START_TIME).total_seconds())
    response = jsonify({
        'status': 'ok',
        'message': 'keepalive',
        'service': 'stock-backend',
        'time': datetime.now().isoformat(),
        'uptime_seconds': uptime_seconds,
        'db_connected': db is not None
    })
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    return response, 200


@app.route('/api/consolidation-status', methods=['GET'])
def consolidation_status():
    """Check if consolidated MCAP/PR averages exist in the database."""
    if symbol_aggregates_collection is None:
        return jsonify({
            'ready': False,
            'mcap_count': 0,
            'pr_count': 0,
            'message': 'Database not connected'
        }), 200

    try:
        mcap_count = symbol_aggregates_collection.count_documents({'type': 'mcap'})
        pr_count = symbol_aggregates_collection.count_documents({'type': 'pr'})
        
        # Ready only if we have both MCAP and PR averages (at least 100 symbols each)
        ready = mcap_count >= 100 and pr_count >= 100
        
        message = ''
        if ready:
            message = f'✅ Ready: {mcap_count} MCAP and {pr_count} PR symbol averages calculated'
        elif mcap_count < 100 and pr_count < 100:
            message = f'⚠️ Need to export Excel: Only {mcap_count} MCAP and {pr_count} PR averages found (need 100+ each)'
        elif mcap_count < 100:
            message = f'⚠️ MCAP averages incomplete: {mcap_count} found (need 100+)'
        else:
            message = f'⚠️ PR averages incomplete: {pr_count} found (need 100+)'
        
        return jsonify({
            'ready': ready,
            'mcap_count': mcap_count,
            'pr_count': pr_count,
            'message': message
        }), 200
    except Exception as e:
        return jsonify({
            'ready': False,
            'mcap_count': 0,
            'pr_count': 0,
            'message': f'Error checking status: {e}'
        }), 200

def calculate_averages_from_consolidated_data(symbols, start_date=None, end_date=None):
    """
    Calculate averages using the same method as consolidation Excel files.
    This ensures consistency between dashboard and consolidation exports.
    """
    if not symbols:
        return {}
    
    try:
        # Build date list the same way as consolidation
        if start_date and end_date:
            try:
                start_dt = date_parser.parse(start_date)
                end_dt = date_parser.parse(end_date)
                if start_dt > end_dt:
                    return {}
                current = start_dt
                date_iso_list = []
                while current <= end_dt:
                    # Removed weekend skip: Include Saturdays and Sundays as well
                    date_iso_list.append(current.strftime('%Y-%m-%d'))
                    current += timedelta(days=1)
            except:
                return {}
        else:
            # If no date range specified, return empty (same as original behavior)
            return {}
        
        if not date_iso_list:
            return {}
        
        print(f"[calculate_averages_from_consolidated_data] Processing {len(date_iso_list)} dates for {len(symbols)} symbols")
        
        # Use the same consolidation logic for MCAP data
        result = {}
        
        try:
            # Get MCAP averages using consolidation method
            mcap_df, dates_list, avg_col = build_consolidated_from_cache(
                date_iso_list, 'mcap', allow_missing=True, log_fn=None,
                allowed_symbols=set(symbols), symbol_name_map=None
            )
            
            # Extract averages for requested symbols
            for _, row in mcap_df.iterrows():
                symbol = row['Symbol']
                if symbol in symbols:
                    if symbol not in result:
                        result[symbol] = {}
                    result[symbol]['avg_total_market_cap'] = row.get(avg_col)
                    
                    # Calculate free float average if Free Float Market Cap columns exist
                    ff_cols = [col for col in mcap_df.columns if 'free' in col.lower() and 'float' in col.lower()]
                    if ff_cols:
                        # Look for daily free float columns
                        date_cols = [c for c in mcap_df.columns if c not in ['Symbol', 'Company Name', 'Days With Data', avg_col]]
                        # This is tricky - the consolidation doesn't separate free float, so use proportional estimate
                        # For now, use the same as total market cap (this should be improved with actual free float data)
                        result[symbol]['avg_free_float_mcap'] = row.get(avg_col)
            
            print(f"[calculate_averages_from_consolidated_data] Got MCAP averages for {len(result)} symbols")
            
        except Exception as exc:
            print(f"⚠️ Failed to get MCAP averages from consolidated data: {exc}")
        
        try:
            # Get PR averages using consolidation method
            # Create symbol-to-name mapping for PR processing
            mcap_name_map = {}
            if 'mcap_df' in locals():
                mcap_name_map = dict(zip(mcap_df['Symbol'], mcap_df['Company Name']))
            
            pr_name_to_symbol = {v: k for k, v in mcap_name_map.items()} if mcap_name_map else None
            
            pr_df, pr_dates_list, pr_avg_col = build_consolidated_from_cache(
                date_iso_list, 'pr', allow_missing=True, log_fn=None,
                allowed_symbols=set(symbols), symbol_name_map=pr_name_to_symbol
            )
            
            # Extract PR averages for requested symbols
            for _, row in pr_df.iterrows():
                symbol = row['Symbol']
                if symbol in symbols:
                    if symbol not in result:
                        result[symbol] = {}
                    result[symbol]['avg_total_traded_value'] = row.get(pr_avg_col)
            
            print(f"[calculate_averages_from_consolidated_data] Got PR averages for {len([s for s in result.values() if 'avg_total_traded_value' in s])} symbols")
            
        except Exception as exc:
            print(f"⚠️ Failed to get PR averages from consolidated data: {exc}")
        
        # Set impact cost to None (not available in consolidation data)
        for symbol_data in result.values():
            if 'avg_impact_cost' not in symbol_data:
                symbol_data['avg_impact_cost'] = None
        
        print(f"[calculate_averages_from_consolidated_data] Final result: averages for {len(result)} symbols")
        return result
        
    except Exception as exc:
        print(f"⚠️ Failed to calculate averages from consolidated data: {exc}")
        return {}


def calculate_averages_from_db(symbols, start_date=None, end_date=None):
    """
    Calculate averages for impact_cost, free_float_mcap, total_market_cap, and total_traded_value
    from the symbol_metrics_daily collection.
    
    DEPRECATED: Use calculate_averages_from_consolidated_data for consistency with consolidation Excel.
    """
    # For backward compatibility, try consolidation method first
    consolidated_averages = calculate_averages_from_consolidated_data(symbols, start_date, end_date)
    if consolidated_averages:
        return consolidated_averages
    
    # Fallback to original DB method with new schema
    if symbol_metrics_daily_collection is None or not symbols:
        return {}
    
    try:
        # New schema: { symbol: "RELIANCE", daily_data: [{date: "2026-02-02", impact_cost: ..., free_float_mcap: ...}, ...] }
        # Build aggregation pipeline to unwind daily_data array and filter by date
        pipeline = [
            {'$match': {'symbol': {'$in': symbols}}},
            {'$unwind': '$daily_data'},
        ]
        
        # Add date filter if provided
        if start_date or end_date:
            date_filter = {}
            if start_date:
                date_filter['$gte'] = start_date
            if end_date:
                date_filter['$lte'] = end_date
            pipeline.append({'$match': {'daily_data.date': date_filter}})
        
        # Calculate averages
        pipeline.extend([
            {
                '$group': {
                    '_id': '$symbol',
                    'avg_impact_cost': {'$avg': '$daily_data.impact_cost'},
                    'avg_free_float_mcap': {'$avg': '$daily_data.free_float_mcap'},
                    'avg_total_market_cap': {'$avg': '$daily_data.total_market_cap'},
                    'avg_total_traded_value': {'$avg': '$daily_data.total_traded_value'},
                    'count': {'$sum': 1}
                }
            }
        ])
        
        result = {}
        for doc in symbol_metrics_daily_collection.aggregate(pipeline):
            symbol = doc['_id']
            result[symbol] = {
                'avg_impact_cost': doc.get('avg_impact_cost'),
                'avg_free_float_mcap': doc.get('avg_free_float_mcap'),
                'avg_total_market_cap': doc.get('avg_total_market_cap'),
                'avg_total_traded_value': doc.get('avg_total_traded_value'),
                'days_count': doc.get('count', 0)
            }
        
        print(f"[calculate_averages_from_db] Calculated averages for {len(result)} symbols from DB (fallback)")
        return result
    except Exception as exc:
        print(f"⚠️ Failed to calculate averages from DB: {exc}")
        return {}


def save_excel_to_database(excel_path, filename, metadata):
    """Save Excel file to MongoDB"""
    if excel_results_collection is None:
        return None
    
    try:
        # Read Excel file as binary
        with open(excel_path, 'rb') as f:
            excel_binary = Binary(f.read())
        
        # Create document for MongoDB
        document = {
            'filename': filename,
            'file_data': excel_binary,
            'file_size': os.path.getsize(excel_path),
            'created_at': datetime.now(),
            'metadata': metadata,
            'file_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
        
        # Insert into database
        result = excel_results_collection.insert_one(document)
        print(f"✅ Excel file saved to MongoDB with ID: {result.inserted_id}")
        return result.inserted_id
    except Exception as e:
        print(f"⚠️ Error saving Excel to database: {e}")
        return None


def format_dashboard_excel(rows, excel_path, start_date=None, end_date=None):
    """
    Format dashboard Excel with required columns and calculations.
    Uses EXACT SAME averages as consolidation Excel by extracting from consolidation source.
    When date range is provided, replaces dashboard averages with consolidation averages.
    """
    try:
        df = pd.DataFrame(rows)
        
        # STRICT FILTER: Never include 'PERMITTED' symbol in any dashboard result
        if not df.empty and 'symbol' in df.columns:
            df = df[df['symbol'].astype(str).str.strip().str.upper() != 'PERMITTED'].copy()
            
        if df.empty:
            return False

        # Fill missing companyName from symbol_aggregates DB
        if symbol_aggregates_collection is not None and 'symbol' in df.columns and 'companyName' in df.columns:
            missing_mask = df['companyName'].isna() | (df['companyName'].astype(str).str.strip() == '')
            if missing_mask.any():
                missing_syms = df.loc[missing_mask, 'symbol'].tolist()
                _cname_map = {}
                for doc in symbol_aggregates_collection.find({'symbol': {'$in': missing_syms}, 'type': 'mcap'}):
                    if doc.get('company_name'):
                        _cname_map[doc['symbol']] = doc['company_name']
                for sym, cname in _cname_map.items():
                    df.loc[df['symbol'] == sym, 'companyName'] = cname

        # Use EXACT same averages as consolidation Excel by extracting from consolidation source
        if start_date and end_date:
            symbols = df['symbol'].tolist() if 'symbol' in df.columns else []
            if symbols:
                try:
                    start_dt = date_parser.parse(start_date)
                    end_dt = date_parser.parse(end_date)
                    
                    if start_dt <= end_dt:
                        current = start_dt
                        date_iso_list = []
                        while current <= end_dt:
                            # Removed weekend skip: Include Saturdays and Sundays as well
                            date_iso_list.append(current.strftime('%Y-%m-%d'))
                            current += timedelta(days=1)
                        
                        if date_iso_list:
                            consolidation_averages = {}
                            import time
                            f_start = time.perf_counter()
                            print(f"[format_dashboard_excel] ⚡ Consolidating averages for {len(symbols)} symbols across {len(date_iso_list)} dates...")
                            
                            # Get MCAP averages using EXACT same method as consolidation
                            try:
                                mcap_df, dates_list, avg_col = build_consolidated_from_cache(
                                    date_iso_list, 'mcap', allow_missing=True, log_fn=None,
                                    allowed_symbols=set(symbols), symbol_name_map=None
                                )
                                
                                mcap_lookup = {row['Symbol']: {'mcap': row[avg_col], 'company_name': row.get('Company Name', '')} 
                                               for _, row in mcap_df.iterrows()}
                                
                                # Store for dashboard symbols
                                for symbol in symbols:
                                    if symbol in mcap_lookup:
                                        consolidation_averages[symbol] = mcap_lookup[symbol]
                                
                                # Get PR averages using EXACT same method as consolidation
                                try:
                                    mcap_name_map = dict(zip(mcap_df['Symbol'], mcap_df['Company Name']))
                                    pr_name_to_symbol = {v: k for k, v in mcap_name_map.items()}
                                    
                                    pr_df, pr_dates_list, pr_avg_col = build_consolidated_from_cache(
                                        date_iso_list, 'pr', allow_missing=True, log_fn=None,
                                        allowed_symbols=set(symbols), symbol_name_map=pr_name_to_symbol
                                    )
                                    
                                    pr_lookup = {row['Symbol']: row[pr_avg_col] for _, row in pr_df.iterrows()}
                                    
                                    # ... (rest of the loop)
                                    for symbol in symbols:
                                        if symbol in pr_lookup and symbol in consolidation_averages:
                                            consolidation_averages[symbol]['traded_value'] = pr_lookup[symbol]
                                            
                                    print(f"[format_dashboard_excel] ✅ Consolidation took {time.perf_counter() - f_start:.2f}s")
                                            
                                except Exception as exc:
                                    print(f"[format_dashboard_excel] ⚠️ PR consolidation averages failed: {exc}")
                                
                                # Replace values in dataframe
                                for idx, row in df.iterrows():
                                    symbol = row.get('symbol')
                                    if symbol and symbol in consolidation_averages:
                                        cons_data = consolidation_averages[symbol]
                                        if 'mcap' in cons_data:
                                            df.at[idx, 'total_market_cap'] = cons_data['mcap']
                                        if 'traded_value' in cons_data:
                                            df.at[idx, 'total_traded_value'] = cons_data['traded_value']
                                            
                            except Exception as exc:
                                print(f"[format_dashboard_excel] ⚠️ MCAP consolidation averages failed: {exc}")
                except Exception as exc:
                    print(f"[format_dashboard_excel] ⚠️ Date processing failed: {exc}")
        
        # Map listingDate to listing_date for consistency
        if 'listingDate' in df.columns and 'listing_date' not in df.columns:
            df['listing_date'] = df['listingDate']
        
        # Optimized calculation of listing info using Vectorized Pandas operations
        if 'listing_date' in df.columns and not df.empty:
            # Convert to datetime once
            ld_series = pd.to_datetime(df['listing_date'], errors='coerce')
            now = datetime.now()
            
            # Days calculation
            df['number of days from listing'] = (now - ld_series).dt.days
            
            # Formatted strings
            df['Day of Listing'] = ld_series.dt.strftime('%d-%b-%Y')
            
            # Boolean logic for 1m/6m
            months_series = df['number of days from listing'] / 30.44
            df['listed> 6months'] = months_series.apply(lambda x: 'Y' if x >= 6 else 'N')
            df['listed> 1 months'] = months_series.apply(lambda x: 'Y' if x >= 1 else 'N')
        else:
            for col in ['Day of Listing', 'listed> 6months', 'listed> 1 months', 'number of days from listing']:
                df[col] = None

        # Determine broader index - Strictly using the freshly fetched Nifty CSV data (Index (API))
        qualifying_indices = {'NIFTY 50', 'NIFTY NEXT 50', 'NIFTY MIDCAP 150', 'NIFTY SMALLCAP 250'}
        # Standardize for comparison
        std_qualifying = {i.replace(' ', '').upper() for i in qualifying_indices}
        
        def is_broader_index(idx):
            if not idx or str(idx).strip().upper() == 'PERMITTED': return ''
            idx_up = str(idx).replace(' ', '').upper()
            return 'NIFTY 500' if idx_up in std_qualifying else ''
            
        df['Broader Index'] = df['index'].apply(is_broader_index)
        
        # Clean "Permitted" from Index (API) column
        if 'index' in df.columns:
            df['index'] = df['index'].apply(lambda x: None if str(x).strip().upper() == 'PERMITTED' else x)
        
        # NOTE: Monetary normalization to Crores moved to the end of processing 
        # to ensure ratio calculations use unit-safe raw values.
        
        if 'impact_cost' in df.columns:
            df['impact_cost'] = pd.to_numeric(df['impact_cost'], errors='coerce').round(2)


        # ── FF/MC Ratio ──────────────────────────────────────────────────────────
        # Use raw units for ratio calculation to ensure unit safety.
        # Safely read live detail columns (may not exist in cached/old rows)
        if 'live_detail_mc' in df.columns:
            df['live_detail_mc_num'] = pd.to_numeric(df['live_detail_mc'], errors='coerce')
        else:
            df['live_detail_mc_num'] = np.nan

        if 'live_detail_ff' in df.columns:
            df['live_detail_ff_num'] = pd.to_numeric(df['live_detail_ff'], errors='coerce')
        else:
            df['live_detail_ff_num'] = np.nan

        # 1. FF/MC from Detail API (raw units → ratio is unit-safe)
        df['FF/MC_Detail'] = (
            df['live_detail_ff_num'] / df['live_detail_mc_num'].replace(0, np.nan)
        ).round(4)

        # 2. Fallback: use average columns (still in raw Rupees here)
        _ff_col = pd.to_numeric(df['free_float_mcap'], errors='coerce') if 'free_float_mcap' in df.columns else pd.Series(np.nan, index=df.index)
        _mc_col = pd.to_numeric(df['total_market_cap'], errors='coerce') if 'total_market_cap' in df.columns else pd.Series(np.nan, index=df.index)
        df['FF/MC_Fallback'] = (_ff_col / _mc_col.replace(0, np.nan)).round(4)

        # Final FF/MC: Detail API first, then fallback
        df['FF/MC'] = df['FF/MC_Detail'].fillna(df['FF/MC_Fallback']).fillna(0)
        
        # REQUIREMENT: Remove stocks where FF/MC ratio is 0.000
        initial_count = len(df)
        df = df[df['FF/MC'] > 0].copy()
        dropped = initial_count - len(df)
        if dropped > 0:
            print(f"[format_excel] Removed {dropped} rows with FF/MC ratio <= 0.000")

        df['Ratio of avg free float to avg total market cap'] = df['FF/MC']
        df['ratio of free float to avg total market cap'] = df['FF/MC']
        
        # 3. Scale primary metrics to Crores AFTER ratio calculations
        monetary_cols = ['total_market_cap', 'free_float_mcap', 'total_traded_value']
        for col in monetary_cols:
            if col in df.columns:
                # Robust scaling: only divide if values appear to be in raw Rupees (> 10M threshold for MCAP)
                # This prevents double-scaling if data was already in Crores.
                vals = pd.to_numeric(df[col], errors='coerce')
                # If median is > 100,000, it's almost certainly raw Rupees (NSE listed companies are valued millions+)
                if vals.median() > 100000:
                    df[col] = (vals / 10000000).round(2)
                else:
                    df[col] = vals.round(2)

        # 4. Calculate recalculate Free Float Market Cap (Cr) using: FF/MC * Current MC (Cr)
        # Use live Detail MC if available, else fallback to aggregated MC.
        live_detail_mc_cr = (df['live_detail_mc_num'] / 10000000).round(2)
        live_mc_num = pd.to_numeric(df['live_mc'], errors='coerce') if 'live_mc' in df.columns else np.nan
        
        mc_source = live_detail_mc_cr.fillna(live_mc_num / 10000000) \
                                     .fillna(df['total_market_cap'] if 'total_market_cap' in df.columns else np.nan)
        
        new_ff_mcap = (df['FF/MC'] * mc_source).round(2)
        if 'free_float_mcap' in df.columns:
            df['free_float_mcap'] = new_ff_mcap.fillna(df['free_float_mcap'])
        else:
            df['free_float_mcap'] = new_ff_mcap

        if 'impact_cost' in df.columns:
            df['impact_cost'] = pd.to_numeric(df['impact_cost'], errors='coerce').round(2)

        # 5. Calculate Trading Consistency (% of Days Traded)
        # Override: Forced to 100% for all symbols as requested
        df['trading_consistency'] = 100.0
        
        # Reorder and rename columns
        output_columns = [
            ('Serial No', 'serial_no'),
            ('Symbol', 'symbol'),
            ('Company name', 'companyName'),
            ('Index (API)', 'index'),
            ('Avg Impact cost', 'impact_cost'),
            ('Average Market Cap (Cr)', 'total_market_cap'),
            ('Average Free Float Market Cap (Cr)', 'free_float_mcap'),
            ('Average Net Traded Value (Cr)', 'total_traded_value'),
            ('% of Days Traded', 'trading_consistency'),
            ('Day of Listing', 'Day of Listing'),
            ('number of days from listing', 'number of days from listing'),
            ('Broader Index', 'Broader Index'),
            ('listed> 6months', 'listed> 6months'),
            ('listed> 1 months', 'listed> 1 months'),
            ('FF/MC', 'FF/MC')
        ]
        
        df['serial_no'] = range(1, len(df) + 1)
        output_df = pd.DataFrame()
        for new_name, old_name in output_columns:
            output_df[new_name] = df[old_name] if old_name in df.columns else None
        
        # HIGH-PERFORMANCE writing with XlsxWriter
        print(f"[format_dashboard_excel] 🚀 Writing {len(output_df)} rows via XlsxWriter...")
        writer = pd.ExcelWriter(excel_path, engine='xlsxwriter')
        output_df.to_excel(writer, index=False, sheet_name='Dashboard')
        
        workbook = writer.book
        worksheet = writer.sheets['Dashboard']
        
        # Define formats
        header_format = workbook.add_format({
            'bold': True, 'font_color': 'white', 'bg_color': '#4472C4',
            'border': 1, 'border_color': '#000000',
            'align': 'center', 'valign': 'vcenter', 'text_wrap': True
        })
        
        cell_format = workbook.add_format({'border': 1, 'border_color': '#000000'})
        num_format = workbook.add_format({'border': 1, 'border_color': '#000000', 'num_format': '#,##0.00'})
        ratio_format = workbook.add_format({'border': 1, 'border_color': '#000000', 'num_format': '0.0000'})
        
        # Column widths & Default Cell Format (Grid)
        # 15 cols (0-14): SerialNo Symbol CompanyName Index ImpactCost AvgMC AvgFF AvgTV
        #                  %Traded DayOfListing DaysFromListing BroaderIdx >6m >1m FF/MC
        widths = [8, 12, 30, 20, 15, 24, 28, 26, 14, 15, 25, 15, 13, 13, 18]
        for i, w in enumerate(widths):
            # Apply cell_format as the default for the column to ensure grid
            worksheet.set_column(i, i, w, cell_format)
            
        # Specific numeric formats for columns
        worksheet.set_column(4, 4, 15, num_format)    # Avg Impact Cost
        worksheet.set_column(5, 5, 24, num_format)    # Avg Market Cap
        worksheet.set_column(6, 6, 28, num_format)    # Avg Free Float
        worksheet.set_column(7, 7, 26, num_format)    # Avg Net Traded Value
        worksheet.set_column(8, 8, 14, num_format)    # % of Days Traded
        worksheet.set_column(14, 14, 18, ratio_format) # FF/MC ratio
        
        # Re-apply header formatting
        for col_num, value in enumerate(output_df.columns.values):
            worksheet.write(0, col_num, value, header_format)
            
        # Define standard Excel light-red format for 'N' values
        warning_format = workbook.add_format({
            'bg_color':   '#FFC7CE',
            'font_color': '#9C0006'
        })
        
        # Apply conditional formatting to 'listed> 6months' (Col 12) and 'listed> 1 months' (Col 13)
        last_row = len(output_df)
        if last_row > 0:
            worksheet.conditional_format(1, 12, last_row, 13, {
                'type':     'cell',
                'criteria': 'equal to',
                'value':    '"N"',
                'format':   warning_format
            })
            
        # Freeze panes
        worksheet.freeze_panes(1, 0)
        
        writer.close()
        
        print(f"[format_dashboard_excel] ✅ High-performance generation took {time.perf_counter() - f_start:.2f}s total")
        return True
    except Exception as e:
        print(f"⚠️ Error formatting dashboard Excel: {e}")
        import traceback
        traceback.print_exc()
        return False
    except Exception as e:
        print(f"⚠️ Error formatting dashboard Excel: {e}")
        import traceback
        traceback.print_exc()
        return False



def download_nse_csv(date_obj_or_iso, data_type):
    """
    Download a specific CSV (mcap or pr) from NSE for a given date.
    Returns: pandas.DataFrame or None
    """
    try:
        if isinstance(date_obj_or_iso, str):
            # Handle both YYYY-MM-DD and DD-Mon-YYYY if needed, but primary is ISO
            try:
                date_obj = datetime.strptime(date_obj_or_iso, '%Y-%m-%d')
            except:
                date_obj = date_parser.parse(date_obj_or_iso)
        else:
            date_obj = date_obj_or_iso
            
        nse_date_formatted = date_obj.strftime('%d-%b-%Y')
        
        # NSE API request
        api_url = "https://www.nseindia.com/api/reports"
        params = {
            'archives': json.dumps([{
                "name": "CM - Bhavcopy (PR.zip)",
                "type": "archives",
                "category": "capital-market",
                "section": "equities"
            }]),
            'date': nse_date_formatted,
            'type': 'equities',
            'mode': 'single'
        }
        
        headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
            'Accept': 'application/json, text/plain, */*',
            'Accept-Language': 'en-US,en;q=0.9',
            'Connection': 'keep-alive'
        }
        
        # Mimic browser requests more closely to avoid blocks
        response = requests.get(api_url, params=params, headers=headers, timeout=30)
        if response.status_code != 200:
            print(f"⚠️ NSE API returned {response.status_code} for {nse_date_formatted}")
            return None
            
        zip_data = BytesIO(response.content)
        with zipfile.ZipFile(zip_data, 'r') as zip_ref:
            file_list = zip_ref.namelist()
            
            target_file = None
            if data_type == 'mcap':
                # Look for mcap or bhav fallback
                for file in file_list:
                    if file.lower().startswith('mcap') and file.lower().endswith('.csv'):
                        target_file = file
                        break
                if not target_file:
                    for file in file_list:
                        if file.lower().endswith('.csv') and ('bhav' in file.lower() or file.lower().startswith('bh')):
                            target_file = file
                            break
            else: # pr
                for file in file_list:
                    if file.lower().startswith('pr') and file.lower().endswith('.csv'):
                        target_file = file
                        break
                        
            if not target_file:
                print(f"⚠️ Target file {data_type} not found in ZIP for {nse_date_formatted}. Files: {file_list}")
                return None
                
            csv_content = zip_ref.read(target_file)
            df = pd.read_csv(BytesIO(csv_content), on_bad_lines='skip', engine='python')
            df.columns = df.columns.str.strip()
            return df
    except Exception as e:
        print(f"⚠️ Error downloading {data_type} for {date_obj_or_iso}: {e}")
        return None


@app.route('/api/download-nse', methods=['POST'])
def download_nse_data():
    """
    Download market cap data from NSE website
    Expected payload: {"date": "03-Dec-2025"}
    Caches CSVs to Mongo and returns metadata (no sessions).
    """
    try:
        data = request.get_json()
        nse_date = data.get('date', '')  # Format: "03-Dec-2025"
        
        if not nse_date:
            return jsonify({'error': 'Date is required in format DD-Mon-YYYY'}), 400
        
        # Parse date
        try:
            date_obj = date_parser.parse(nse_date)
            date_iso = date_obj.strftime('%Y-%m-%d')
            nse_date_formatted = date_obj.strftime('%d-%b-%Y')
        except:
            return jsonify({'error': 'Invalid date format. Use DD-Mon-YYYY (e.g., 03-Dec-2025)'}), 400
        
        print(f"Downloading NSE data for {nse_date_formatted}...")
        
        # Use helper for MCAP
        mcap_df = download_nse_csv(date_obj, 'mcap')
        if mcap_df is None:
            return jsonify({'error': f'Failed to download or parse MCAP for {nse_date_formatted}'}), 404
            
        # Use helper for PR
        pr_df = download_nse_csv(date_obj, 'pr')
        
        # Persist to Mongo cache and symbol_daily
        put_cached_csv(date_iso, 'mcap', mcap_df, source='nse')
        bulk_upsert_symbol_daily_from_df(mcap_df, date_iso, 'mcap', source='nse_download')
        
        if pr_df is not None:
            put_cached_csv(date_iso, 'pr', pr_df, source='nse')
            bulk_upsert_symbol_daily_from_df(pr_df, date_iso, 'pr', source='nse_download')
        
        return jsonify({
            'success': True,
            'message': 'Files downloaded and cached to Mongo',
            'files': {
                'mcap': {
                    'records': len(mcap_df),
                    'columns': mcap_df.columns.tolist()
                },
                'pr': {
                    'records': len(pr_df) if pr_df is not None else 0,
                    'columns': pr_df.columns.tolist() if pr_df is not None else []
                }
            },
            'date': nse_date_formatted
        }), 200
    
    except Exception as e:
        print(f"Error in download_nse_data: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/nse-dates', methods=['GET'])
def get_nse_dates():
    """
    Get available dates for NSE data (last 2 years of trading days)
    Returns list of dates in DD-Mon-YYYY format
    """
    try:
        dates = []
        today = datetime.now()
        
        # Generate dates for last 2 years (approximately 500+ trading days, excluding weekends)
        # 2 years = ~730 days, ~500 trading days (excluding weekends/holidays)
        start_date = today - timedelta(days=730)
        
        current = start_date
        while current <= today:
            # Removed weekend skip: Include Saturdays and Sundays as well
            dates.append(current.strftime('%d-%b-%Y'))
            current += timedelta(days=1)
        
        # Reverse to show most recent first
        dates.reverse()
        
        return jsonify({
            'success': True,
            'dates': dates,
            'today': today.strftime('%d-%b-%Y'),
            'count': len(dates)
        }), 200
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/download-nse-range', methods=['POST'])
def download_nse_range():
    """
    Download market cap data for a date range from NSE website
    Expected payload: {
        "start_date": "01-Dec-2025",  # Format: DD-Mon-YYYY
        "end_date": "05-Dec-2025",    # Format: DD-Mon-YYYY
        "refresh_mode": "missing_only" or "force"  # Optional, default missing_only
    }
    Returns summary and entries (no sessions; files cached in Mongo).
    """
    try:
        data = request.get_json()
        start_date_str = data.get('start_date', '')
        end_date_str = data.get('end_date', '')
        refresh_mode = data.get('refresh_mode', 'missing_only')  # missing_only | force
        parallel_workers = int(data.get('parallel_workers', 20) or 20)

        if refresh_mode not in ['missing_only', 'force']:
            return jsonify({'error': 'Invalid refresh_mode. Use missing_only or force'}), 400

        if not start_date_str or not end_date_str:
            return jsonify({'error': 'Both start_date and end_date are required'}), 400
        
        # Parse dates
        try:
            start_date = date_parser.parse(start_date_str)
            end_date = date_parser.parse(end_date_str)
        except:
            return jsonify({'error': 'Invalid date format. Use DD-Mon-YYYY (e.g., 01-Dec-2025)'}), 400
        
        if start_date > end_date:
            return jsonify({'error': 'start_date cannot be after end_date'}), 400
        
        # Generate list of trading days in range
        current_date = start_date
        trading_dates = []
        
        while current_date <= end_date:
            # Removed weekend skip: Include Saturdays and Sundays as well
            trading_dates.append(current_date)
            current_date += timedelta(days=1)
        
        if not trading_dates:
            return jsonify({'error': 'No trading days found in the selected range'}), 400
        
        downloads_summary = {
            'total_requested': len(trading_dates),
            'cached_count': 0,
            'fetched_count': 0,
            'failed_count': 0,
            'entries': [],
            'errors': []
        }

        print(f"Downloading NSE data for {len(trading_dates)} trading days (mode={refresh_mode}, workers={parallel_workers})...")

        # OPTIMIZATION: Bulk fetch existing cache METADATA ONLY first
        # This is extremely fast because it excludes the heavy file_data blobs.
        date_iso_list = [_normalize_iso_date(d) for d in trading_dates]
        prefetched_mcap_meta = {}
        prefetched_pr_meta = {}
        if refresh_mode != 'force':
            print(f"[download-range] Fast-checking cache metadata for {len(date_iso_list)} dates...")
            prefetched_mcap_meta = get_cached_csv_metadata_bulk(date_iso_list, 'mcap')
            prefetched_pr_meta = get_cached_csv_metadata_bulk(date_iso_list, 'pr')

        def process_trade_date(index, trade_date):
            nse_date_formatted = trade_date.strftime('%d-%b-%Y')
            date_iso = _normalize_iso_date(trade_date)

            result_entry = {
                'index': index + 1,
                'date': nse_date_formatted,
                'status': 'fetched',
                'pr_status': 'fetched',
                'mcap_records': 0,
                'pr_records': 0
            }

            # 1. LIGHTWEIGHT CACHE CHECK
            mcap_meta = prefetched_mcap_meta.get(date_iso) if refresh_mode != 'force' else None
            pr_meta = prefetched_pr_meta.get(date_iso) if refresh_mode != 'force' else None
            # 2. FAST PATH: If both are cached and NOT force mode, return IMMEDIATELY
            if mcap_meta and pr_meta and refresh_mode != 'force':
                result_entry['status'] = 'cached'
                result_entry['pr_status'] = 'cached'
                result_entry['mcap_records'] = mcap_meta.get('records', 0)
                result_entry['pr_records'] = pr_meta.get('records', 0)
                return index, {
                    'entries': [result_entry],
                    'errors': [],
                    'cached_count': 1,
                    'fetched_count': 0,
                    'failed_count': 0
                }

            # 3. SLOW PATH: Missing data or FORCE mode
            # We need the actual dataframes for these cases
            try:
                # MCAP Stage
                mcap_df = None
                if mcap_meta and refresh_mode != 'force':
                    result_entry['status'] = 'cached'
                    mcap_records = mcap_meta.get('records', 0)
                else:
                    # Actually fetch from NSE or full cache
                    mcap_data = get_cached_csv(date_iso, 'mcap') if refresh_mode != 'force' else None
                    if not mcap_data:
                        mcap_df = download_nse_csv(trade_date, 'mcap')
                        if mcap_df is not None:
                            put_cached_csv(date_iso, 'mcap', mcap_df)
                    else:
                        mcap_df = mcap_data['df']
                    
                    if mcap_df is not None:
                        bulk_upsert_symbol_daily_from_df(mcap_df, date_iso, 'mcap', source='nse_download')
                        mcap_records = len(mcap_df)
                    else:
                        mcap_records = 0
                
                result_entry['mcap_records'] = mcap_records

                # PR Stage
                pr_df = None
                if pr_meta and refresh_mode != 'force':
                    result_entry['pr_status'] = 'cached'
                    pr_records = pr_meta.get('records', 0)
                else:
                    pr_data = get_cached_csv(date_iso, 'pr') if refresh_mode != 'force' else None
                    if not pr_data:
                        pr_df = download_nse_csv(trade_date, 'pr')
                        if pr_df is not None:
                            put_cached_csv(date_iso, 'pr', pr_df)
                    else:
                        pr_df = pr_data['df']
                    
                    if pr_df is not None:
                        # Build name map from MCAP for PR symbol mapping
                        symbol_name_map = {}
                        if mcap_df is not None:
                            symbol_name_map = dict(zip(mcap_df['Security Name'], mcap_df['Symbol']))
                        elif result_entry['status'] == 'cached':
                            # Try to get map from cached MCAP
                            mcap_data = get_cached_csv(date_iso, 'mcap')
                            if mcap_data:
                                symbol_name_map = dict(zip(mcap_data['df']['Security Name'], mcap_data['df']['Symbol']))
                                
                        bulk_upsert_symbol_daily_from_df(pr_df, date_iso, 'pr', source='nse_download', symbol_name_map=symbol_name_map)
                        pr_records = len(pr_df)
                    else:
                        pr_records = 0
                
                result_entry['pr_records'] = pr_records

                is_cached = (result_entry['status'] == 'cached' and result_entry['pr_status'] == 'cached')
                
                return index, {
                    'entries': [result_entry],
                    'errors': [],
                    'cached_count': 1 if is_cached else 0,
                    'fetched_count': 0 if is_cached else 1,
                    'failed_count': 0
                }

            except Exception as e:
                print(f"❌ Error processing {date_iso}: {e}")
                return index, {
                    'entries': [result_entry],
                    'errors': [{'date': nse_date_formatted, 'error': str(e)}],
                    'cached_count': 0,
                    'fetched_count': 0,
                    'failed_count': 1
                }

        futures = {}
        with ThreadPoolExecutor(max_workers=parallel_workers) as executor:
            for idx, trade_date in enumerate(trading_dates):
                futures[executor.submit(process_trade_date, idx, trade_date)] = idx

            results_by_index = {}
            for future in as_completed(futures):
                idx = futures[future]
                try:
                    idx, res = future.result()
                    results_by_index[idx] = res
                except Exception as exc:
                    results_by_index[idx] = {
                        'entries': [],
                        'errors': [{
                            'date': trading_dates[idx].strftime('%d-%b-%Y'),
                            'error': f'Worker error: {exc}'
                        }],
                        'cached_count': 0,
                        'fetched_count': 0,
                        'failed_count': 1
                    }

        for idx in range(len(trading_dates)):
            res = results_by_index.get(idx, None)
            if res is None:
                continue
            downloads_summary['cached_count'] += res['cached_count']
            downloads_summary['fetched_count'] += res['fetched_count']
            downloads_summary['failed_count'] += res['failed_count']
            downloads_summary['entries'].extend(res['entries'])
            downloads_summary['errors'].extend(res['errors'])

        # Categorize errors for better user feedback
        error_categories = {
            'holidays': [],
            'network': [],
            'other': []
        }
        
        for error in downloads_summary['errors']:
            error_msg = error.get('error', '').lower()
            if '404' in error_msg or 'no data available' in error_msg or 'holiday' in error_msg:
                error_categories['holidays'].append(error['date'])
            elif 'connection' in error_msg or 'resolve' in error_msg or 'getaddrinfo' in error_msg or 'timeout' in error_msg:
                error_categories['network'].append(error['date'])
            else:
                error_categories['other'].append(error)
        
        # Build helpful error summary
        error_summary = []
        if error_categories['holidays']:
            error_summary.append(f"No data available (likely holidays): {', '.join(error_categories['holidays'])}")
        if error_categories['network']:
            error_summary.append(f"Network/connection errors: {', '.join(error_categories['network'])} - Try again later or check internet connection")
        if error_categories['other']:
            error_summary.append(f"{len(error_categories['other'])} other errors - check details below")

        return jsonify({
            'success': True,
            'summary': {
                'total_requested': downloads_summary['total_requested'],
                'cached': downloads_summary['cached_count'],
                'fetched': downloads_summary['fetched_count'],
                'failed': downloads_summary['failed_count'],
                'refresh_mode': refresh_mode,
                'parallel_workers': parallel_workers,
                'error_summary': error_summary
            },
            'entries': downloads_summary['entries'],
            'errors': downloads_summary['errors']
        }), 200
    
    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/nse-symbol-dashboard', methods=['POST'])
def nse_symbol_dashboard():
    """
    Build a dashboard of impact cost, free float market cap, traded value, and index for symbols.
    Optimized for Render free tier (30s timeout).
    Uses batch processing from frontend to avoid network errors.
    """
    try:
        data = request.get_json() or {}
        req_id = f"symboldash-{int(time.time() * 1000)}"
        start_time = time.perf_counter()
        
        batch_index = data.get('batch_index')
        date_str = data.get('date')
        start_date_str = data.get('start_date')
        end_date_str = data.get('end_date')
        provided_symbols = data.get('symbols') or []
        top_n = data.get('top_n', 100)
        as_on = datetime.now().strftime('%Y-%m-%d')

        TOTAL_SYMBOLS = 1100
        BATCH_SIZE = 100

        symbols_to_process = provided_symbols[:] if provided_symbols else []
        
        if not symbols_to_process:
            if symbol_aggregates_collection is not None:
                db_symbols = list(symbol_aggregates_collection.find(
                    {'type': 'mcap'},
                    {'symbol': 1, 'average': 1}
                ).sort([('average', -1), ('symbol', 1)]).limit(TOTAL_SYMBOLS))
                if db_symbols:
                    # STRICT FILTER: Never include 'PERMITTED' 
                    symbols_to_process = [s['symbol'] for s in db_symbols if str(s.get('symbol')).strip().upper() != 'PERMITTED']
        
        # GLOBAL FILTER: Ensure 'PERMITTED' is removed from any provided list as well
        if symbols_to_process:
            symbols_to_process = [s for s in symbols_to_process if str(s).strip().upper() != 'PERMITTED']

        if not symbols_to_process:
            market_cap_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'nosubject', 'Market_Cap.xlsx')
            if os.path.exists(market_cap_path):
                df = pd.read_excel(market_cap_path)
                sym_col = next((c for c in df.columns if str(c).strip().lower() in ['symbol', 'symbols']), None)
                if sym_col:
                    symbols_to_process = df[sym_col].astype(str).tolist()[:TOTAL_SYMBOLS]

        if not symbols_to_process:
            return jsonify({'error': 'No symbols found'}), 400

        final_symbols = list(dict.fromkeys(symbols_to_process))[:TOTAL_SYMBOLS]
        for rs in ['GANECOS', 'ALLCARGO']:
            if rs not in final_symbols: final_symbols.append(rs)

        if nifty_indices_collection is not None:
            for _sym in ['GANECOS', 'ALLCARGO']:
                nifty_indices_collection.update_one({'symbol': _sym}, {'$set': {'symbol': _sym, 'indices': ['NIFTY MICROCAP 250'], 'primary_index': 'NIFTY MICROCAP 250', 'last_updated': datetime.now()}}, upsert=True)

        total_count = len(final_symbols)
        if batch_index is None:
            batch_symbols = final_symbols
            batch_idx, total_batches = 0, 1
        else:
            batch_idx = int(batch_index)
            symbol_batches = [final_symbols[i:i + BATCH_SIZE] for i in range(0, len(final_symbols), BATCH_SIZE)]
            total_batches = len(symbol_batches)
            if batch_idx >= total_batches:
                return jsonify({'success': True, 'complete': True, 'rows': []}), 200
            batch_symbols = symbol_batches[batch_idx]

        symbol_pr_data, symbol_mcap_data, index_mapping, metrics_cache = {}, {}, {}, {}

        if symbol_aggregates_collection is not None:
            for doc in symbol_aggregates_collection.find({'symbol': {'$in': batch_symbols}}):
                sym = str(doc.get('symbol') or '').strip().upper()
                dtype = doc.get('type')
                if not sym: continue
                
                if dtype == 'pr':
                    # Robust key check for legacy/new naming
                    nz_days = doc.get('non_zero_days') if doc.get('non_zero_days') is not None else doc.get('Non Zero Days', 0)
                    symbol_pr_data[sym] = {
                        'days_with_data': doc.get('days_with_data', 0), 
                        'non_zero_days': nz_days,
                        'total_possible_days': doc.get('total_possible_days', 0),
                        'avg_pr': doc.get('average')
                    }
                    if sym not in symbol_mcap_data: symbol_mcap_data[sym] = {}
                    symbol_mcap_data[sym]['total_traded_value'] = doc.get('average')
                    # Propagate consistency fields to the main map used by fetcher
                    symbol_mcap_data[sym]['non_zero_days'] = nz_days
                    symbol_mcap_data[sym]['total_possible_days'] = doc.get('total_possible_days', 0)
                elif dtype == 'mcap':
                    if sym not in symbol_mcap_data: symbol_mcap_data[sym] = {}
                    symbol_mcap_data[sym]['avg_mcap'] = doc.get('average')
                    symbol_mcap_data[sym]['total_possible_days'] = doc.get('total_possible_days', 0)
                    # Robust key check for legacy/new naming
                    symbol_mcap_data[sym]['non_zero_days'] = doc.get('non_zero_days') if doc.get('non_zero_days') is not None else doc.get('Non Zero Days', 0)

        if nifty_indices_collection is not None:
            for doc in nifty_indices_collection.find({'symbol': {'$in': batch_symbols}}):
                sym = str(doc.get('symbol') or '').strip().upper()
                if sym and doc.get('indices'):
                    index_mapping[sym] = doc['indices']
                    # Inject live data from MongoDB Constituents (if available)
                    if sym not in symbol_mcap_data: symbol_mcap_data[sym] = {}
                    symbol_mcap_data[sym]['live_mc'] = doc.get('live_mc')
                    symbol_mcap_data[sym]['live_ff'] = doc.get('live_ff')

        if symbol_metrics_collection is not None:
            for doc in symbol_metrics_collection.find({'as_on': as_on, 'symbol': {'$in': batch_symbols}}):
                sym = str(doc.get('symbol') or '').strip().upper()
                if sym: metrics_cache[sym] = doc

        fetcher = SymbolMetricsFetcher()
        result = fetcher.build_dashboard(
            batch_symbols, as_of=as_on, parallel=True, max_workers=50, 
            chunk_size=len(batch_symbols),  # Single batch = all symbols processed in parallel
            symbol_pr_data=symbol_pr_data, symbol_mcap_data=symbol_mcap_data,
            external_index_mapping=index_mapping, external_metrics_cache=metrics_cache
        )
        
        rows = result.get('rows', [])
        errors = result.get('errors', [])

        index_map_detailed = primary_index_map_from_db(batch_symbols)
        for row in rows:
            row.pop('_id', None)
            sym = str(row.get('symbol') or '').strip().upper()
            if sym in index_map_detailed: 
                row['primary_index'] = index_map_detailed[sym]
            else:
                row['primary_index'] = None  # EXPLICITLY clear stale index tags
            
            # Inject metrics from DB aggregates
            if sym in symbol_pr_data:
                row['days_with_data'] = symbol_pr_data[sym].get('days_with_data', 0)
                row['non_zero_days'] = symbol_pr_data[sym].get('non_zero_days', 0)
                row['total_possible_days'] = symbol_pr_data[sym].get('total_possible_days', 0)
            elif sym in symbol_mcap_data:
                row['non_zero_days'] = symbol_mcap_data[sym].get('non_zero_days', 0)
                row['total_possible_days'] = symbol_mcap_data[sym].get('total_possible_days', 0)
        
        # High-performance bulk upsert
        bulk_upsert_symbol_metrics(rows, source='symbol_dashboard')

        is_last = (batch_idx == total_batches - 1)
        
        final_data = {
            'success': True, 'count': len(rows), 'rows': rows, 'errors': errors,
            'batch_index': batch_idx, 'total_batches': total_batches,
            'total_symbols': total_count, 'complete': is_last,
            'message': f'Dashboard batch {batch_idx + 1}/{total_batches} complete!', 'percentage': 100
        }
        
        return jsonify(final_data), 200

    except Exception as e:
        print(f"[symbol-dashboard][error] {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


def process_symbol_batch(symbols, as_on, max_workers, timeout, symbol_pr_data=None, symbol_mcap_data=None):
    """Process a batch of symbols with 10 workers for parallel processing."""
    fetcher = SymbolMetricsFetcher()
    try:
        # Use exactly 10 workers for optimal parallel processing
        effective_workers = 10
        # Larger chunk size so each worker handles more symbols at once
        chunk_size = max(10, len(symbols) // 10)  # Divide work among 10 workers
        print(f"[batch-processing] Fetching {len(symbols)} symbols with {effective_workers} workers, chunk_size={chunk_size}")
        
        # Filter to only the symbols in this batch
        batch_pr_data = {s: symbol_pr_data.get(s) for s in symbols if symbol_pr_data and s in symbol_pr_data} if symbol_pr_data else {}
        batch_mcap_data = {s: symbol_mcap_data.get(s) for s in symbols if symbol_mcap_data and s in symbol_mcap_data} if symbol_mcap_data else {}
        
        result = fetcher.build_dashboard(
            symbols,
            excel_path=None,
            max_symbols=None,
            as_of=as_on,
            parallel=True,
            max_workers=effective_workers,
            chunk_size=chunk_size,
            symbol_pr_data=batch_pr_data,
            symbol_mcap_data=batch_mcap_data,
            max_time_seconds=timeout,
            fetch_indices_from_csv=False,
            nifty_indices_collection=nifty_indices_collection
        )
        return result.get('rows', []), result.get('errors', [])
    except Exception as exc:
        return [], [{'error': str(exc)}]


@app.route('/api/nse-symbol-dashboard/download', methods=['GET'])
def download_symbol_dashboard_file():
    if excel_results_collection is None:
        return jsonify({'error': 'Database not connected'}), 500

    file_id = request.args.get('id')
    if not file_id:
        return jsonify({'error': 'Missing id parameter'}), 400

    try:
        oid = ObjectId(file_id)
    except Exception:
        return jsonify({'error': 'Invalid file id'}), 400

    try:
        doc = excel_results_collection.find_one({'_id': oid})
        if not doc:
            return jsonify({'error': 'File not found'}), 404

        return send_file(
            BytesIO(doc['file_data']),
            mimetype=doc.get('file_type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
            as_attachment=True,
            download_name=doc.get('filename', 'Symbol_Dashboard.xlsx')
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/nse-symbol-dashboard/save-excel', methods=['POST'])
def nse_symbol_dashboard_save_excel():
    """
    Accepts all dashboard rows from the frontend and generates a single Excel file for all batches.
    Returns the download_url for the complete file.
    """
    try:
        data = request.get_json() or {}
        rows = data.get('rows', [])
        as_on = data.get('as_on') or datetime.now().strftime('%Y-%m-%d')
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        
        if not rows or not isinstance(rows, list):
            return jsonify({'error': 'No rows provided'}), 400
            
        download_name = f"Symbol_Dashboard_All_{len(rows)}.xlsx"
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        excel_path = temp_file.name
        temp_file.close()
        
        # Use formatted Excel output with date range
        format_dashboard_excel(rows, excel_path, start_date=start_date, end_date=end_date)
        
        db_id = save_excel_to_database(excel_path, download_name, {
            'symbols': len(rows),
            'as_on': as_on,
            'all_batches': True,
            'start_date': start_date,
            'end_date': end_date
        })
        
        download_url = None
        if db_id:
            download_url = f"/api/nse-symbol-dashboard/download?id={db_id}"
        
        os.remove(excel_path)
        
        return jsonify({
            'success': True,
            'file': download_name,
            'file_id': str(db_id) if db_id else None,
            'download_url': download_url,
            'symbols': len(rows)
        }), 200
    except Exception as e:
        print(f"[symbol-dashboard][save-excel][error] {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/consolidate-saved', methods=['POST'])
def consolidate_saved():
    """
    Memory-optimized consolidation of cached NSE CSVs from Mongo into Excel.
    
    Payload (JSON):
    {
        "date": "03-Dec-2025"  # optional, single date
        "start_date": "01-Dec-2025", "end_date": "05-Dec-2025"  # optional range
        "file_type": "both" | "mcap" | "pr"  # default both
        "fast_mode": true/false  # default true, skip DB writes when true
        "optimize_memory": true/false  # default true, use memory optimization
        "max_records_per_batch": 5000  # default 5000, process data in batches
    }

    Response: Excel file (zip when both MCAP and PR are produced).
    """
    work_dir = None  # Initialize early to avoid UnboundLocalError in exception handlers
    req_id = None
    try:
        try:
            payload = request.get_json() or {}
            req_id = f"consolidate-{int(time.time() * 1000)}"
            stage_start = time.perf_counter()
            
            # Memory optimization settings
            optimize_memory = payload.get('optimize_memory', True)
            max_records_per_batch = payload.get('max_records_per_batch', 5000)
            
            # Log initial memory usage
            initial_memory = get_memory_usage_mb()
            print(f"[consolidate-saved][start] id={req_id} initial_memory={initial_memory:.1f}MB optimize_memory={optimize_memory}")
            
            date_str = payload.get('date')
            start_date_str = payload.get('start_date')
            end_date_str = payload.get('end_date')
            file_type = payload.get('file_type', 'both')
            fast_mode = payload.get('fast_mode', False)
            skip_daily = payload.get('skip_daily', True)
            allow_missing = payload.get('allow_missing', True)

            if file_type not in ['mcap', 'pr', 'both']:
                return jsonify({'error': 'Invalid file_type (mcap, pr, or both)'}), 400

            # Build date list
            date_iso_list = []
            try:
                if date_str:
                    date_dt = date_parser.parse(date_str)
                    date_iso_list = [date_dt.strftime('%Y-%m-%d')]
                elif start_date_str and end_date_str:
                    start_dt = date_parser.parse(start_date_str)
                    end_dt = date_parser.parse(end_date_str)
                    if start_dt > end_dt:
                        return jsonify({'error': 'start_date cannot be after end_date'}), 400
                    current = start_dt
                    while current <= end_dt:
                        # Removed weekend skip: Include Saturdays and Sundays as well
                        date_iso_list.append(current.strftime('%Y-%m-%d'))
                        current += timedelta(days=1)
                else:
                    return jsonify({'error': 'Provide either date or start_date/end_date'}), 400
            except Exception:
                return jsonify({'error': 'Invalid date format. Use DD-Mon-YYYY (e.g., 03-Dec-2025)'}), 400

            # Check memory limits - if too many dates, force batching
            if len(date_iso_list) > 7 and not optimize_memory:
                optimize_memory = True
                print(f"[consolidate-saved][warning] Processing {len(date_iso_list)} dates - forcing memory optimization")

            logs = []
            work_dir = tempfile.mkdtemp()
            results = {}

            def add_log(message):
                logs.append(message)
                print(message)

            def make_consolidator_from_cache_optimized(data_type, allowed_symbols=None, symbol_name_map=None):
                """Memory-optimized consolidation from cache"""
                memory_before = get_memory_usage_mb()
                
                # IMPORTANT: Don't batch the consolidation itself - that causes data loss!
                # Instead, rely on build_consolidated_from_cache's internal optimizations
                # Only for EXTREMELY large date ranges (6+ months), process all at once
                
                add_log(f"Processing {data_type.upper()} for {len(date_iso_list)} dates")
                
                df, dates_list, avg_col = build_consolidated_from_cache(
                    date_iso_list, data_type, allow_missing=allow_missing, log_fn=add_log,
                    allowed_symbols=allowed_symbols, symbol_name_map=symbol_name_map
                )
                
                if df is None or df.empty:
                    raise ValueError(f"No {data_type.upper()} data available for requested dates")
                
                memory_after = get_memory_usage_mb()
                add_log(f"{data_type.upper()} consolidated: {len(df)} companies across {len(dates_list)} dates")
                add_log(f"Memory usage: {memory_before:.1f}MB -> {memory_after:.1f}MB (Δ{memory_after-memory_before:+.1f}MB)")
                
                cons = MarketCapConsolidator(work_dir, file_type=data_type)
                cons.df_consolidated = df
                cons.dates_list = dates_list
                cons.avg_col = avg_col
                cons.days_col = 'Days With Data'
                return cons, len(df), len(dates_list)

            mcap_output_path = None
            pr_output_path = None
            mcap_symbols = None
            mcap_name_map = None

            # Initialize memory optimizer
            optimizer = MemoryOptimizedExporter(compression_level=9)  # Maximum compression

            # Build a label for filenames based on requested dates
            if len(date_iso_list) == 1:
                date_label = date_iso_list[0]
            elif len(date_iso_list) > 1:
                date_label = f"{date_iso_list[0]}_to_{date_iso_list[-1]}"
            else:
                date_label = "dates"
            date_label = date_label.replace('/', '-').replace(' ', '_')

            # Data collection for multi-sheet Excel
            excel_sheets = {}

            # Pre-fetch Company Name -> Symbol mapping from Mongo to allow parallel processing
            prefetched_name_map = {}
            if symbol_aggregates_collection is not None:
                try:
                    for doc in symbol_aggregates_collection.find({'type': 'mcap'}, {'symbol': 1, 'company_name': 1}):
                        s, c = doc.get('symbol'), doc.get('company_name')
                        if s and c: prefetched_name_map[c] = s
                    if prefetched_name_map:
                        add_log(f"✓ Pre-fetched {len(prefetched_name_map)} symbol mappings from DB")
                except:
                    pass

            def run_mcap():
                try:
                    mcap_start = time.perf_counter()
                    cons, count, d_count = make_consolidator_from_cache_optimized('mcap')
                    
                    # Extract map for PR just in case it's more up-to-date
                    m_map = dict(zip(cons.df_consolidated['Symbol'], cons.df_consolidated['Company Name']))
                    
                    sheet_df = cons.df_consolidated.copy()
                    
                    if not fast_mode:
                        persist_start = time.perf_counter()
                        add_log(f"Starting MCAP persistence to Mongo...")
                        persist_consolidated_results(cons, 'mcap', source='cached_db', skip_daily=skip_daily)
                        add_log(f"✓ Persisted {count} MCAP averages in {time.perf_counter() - persist_start:.2f}s")
                    
                    return {
                        'cons': cons, 'sheet_df': sheet_df, 'm_map': m_map, 
                        'results': {'companies': count, 'dates': d_count, 'files': len(date_iso_list), 'persisted': not fast_mode},
                        'duration': time.perf_counter() - mcap_start
                    }
                except Exception as e:
                    return {'error': f"MCAP failed: {e}"}

            def run_pr(name_map):
                try:
                    pr_start = time.perf_counter()
                    # Use provided name_map (pre-fetched or live)
                    pr_map = {v: k for k, v in name_map.items()} if name_map else None
                    cons, count, d_count = make_consolidator_from_cache_optimized(
                        'pr', allowed_symbols=None, symbol_name_map=pr_map
                    )
                    
                    sheet_df = cons.df_consolidated.copy()
                    
                    if not fast_mode:
                        persist_start = time.perf_counter()
                        add_log(f"Starting PR persistence to Mongo...")
                        persist_consolidated_results(cons, 'pr', source='cached_db', skip_daily=skip_daily)
                        add_log(f"✓ Persisted {count} PR averages in {time.perf_counter() - persist_start:.2f}s")
                    
                    return {
                        'cons': cons, 'sheet_df': sheet_df,
                        'results': {'companies': count, 'dates': d_count, 'files': len(date_iso_list), 'persisted': not fast_mode},
                        'duration': time.perf_counter() - pr_start
                    }
                except Exception as e:
                    return {'error': f"PR failed: {e}"}

            # Parallel Execution
            with ThreadPoolExecutor(max_workers=2) as executor:
                mcap_task = None
                pr_task = None
                
                if file_type in ['mcap', 'both']:
                    mcap_task = executor.submit(run_mcap)
                
                # If only PR, run with prefetched map
                if file_type == 'pr':
                    pr_task = executor.submit(run_pr, prefetched_name_map)
                
                # Wait for MCAP if it's running, then start PR if 'both'
                mcap_data = mcap_task.result() if mcap_task else None
                if mcap_data and 'error' in mcap_data:
                    shutil.rmtree(work_dir, ignore_errors=True)
                    return jsonify({'error': mcap_data['error']}), 500
                
                if mcap_data:
                    excel_sheets['Market_Cap'] = mcap_data['sheet_df']
                    results['mcap'] = mcap_data['results']
                    add_log(f"MCAP stage done in {mcap_data['duration']:.2f}s")
                    
                    # If both, we can now run PR with potentially better map from MCAP
                    if file_type == 'both':
                        # Merge prefetched and live map
                        combined_map = {**prefetched_name_map, **(mcap_data['m_map'] or {})}
                        pr_task = executor.submit(run_pr, combined_map)
                
                # We can now clear MCAP cons from memory if it finished
                if mcap_data and 'cons' in mcap_data:
                    del mcap_data['cons'].df_consolidated
                    del mcap_data['cons']
                    gc.collect()

                pr_data = pr_task.result() if pr_task else None
                if pr_data and 'error' in pr_data:
                    shutil.rmtree(work_dir, ignore_errors=True)
                    return jsonify({'error': pr_data['error']}), 500
                
                if pr_data:
                    excel_sheets['Net_Traded_Value'] = pr_data['sheet_df']
                    results['pr'] = pr_data['results']
                    add_log(f"PR stage done in {pr_data['duration']:.2f}s")
                    
                    # Clear memory
                    if 'cons' in pr_data:
                        del pr_data['cons'].df_consolidated
                        del pr_data['cons']
                    gc.collect()

            if not excel_sheets:
                shutil.rmtree(work_dir, ignore_errors=True)
                return jsonify({'error': 'No data collected for Excel creation'}), 500

            # Scale values to Crores and round to 2 decimal places for consolidation Excel only
            for sheet_name, df_sheet in excel_sheets.items():
                # REQUIREMENT: Hide intermediate count columns from final report
                cols_to_drop = [c for c in ['non_zero_days', 'total_possible_days'] if c in df_sheet.columns]
                if cols_to_drop:
                    df_sheet.drop(columns=cols_to_drop, inplace=True)
                
                # Identify columns that should be numeric (excluding Symbol, Company Name, Days With Data)
                cols_to_scale = [c for c in df_sheet.columns if c not in ['Symbol', 'Company Name', 'Days With Data']]
                for col in cols_to_scale:
                    try:
                        df_sheet[col] = pd.to_numeric(df_sheet[col], errors='coerce')
                        if pd.api.types.is_numeric_dtype(df_sheet[col]):
                            # IMPORTANT: Only scale MCAP and Traded Value, NOT impact cost or ratios
                            # These columns only contain MCAP or Net Traded Value depending on the sheet
                            df_sheet[col] = (df_sheet[col] / 10000000).round(2)
                    except:
                        pass
                add_log(f"✓ Scaled to Crores and cleaned up sheet '{sheet_name}'")

            # Create single Excel file with multiple sheets
            excel_creation_start = time.perf_counter()
            excel_filename = f"Market_Data_{date_label}.xlsx"
            excel_path = os.path.join(work_dir, excel_filename)
            
            add_log(f"Creating multi-sheet Excel file: {excel_filename}")
            add_log(f"Sheets to create: {list(excel_sheets.keys())}")
            
            if optimize_memory:
                optimizer.create_multi_sheet_excel(excel_sheets, excel_path)
            else:
                # Fallback to openpyxl for multiple sheets
                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    for sheet_name, df in excel_sheets.items():
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Get file size and log
            excel_size_mb = os.path.getsize(excel_path) / 1024 / 1024
            add_log(f"✓ Multi-sheet Excel created: {excel_size_mb:.1f}MB in {time.perf_counter() - excel_creation_start:.2f}s")
            
            # Clear sheets data from memory
            del excel_sheets
            gc.collect()

            # Always copy Market_Cap sheet to nosubject/ if MCAP data exists
            try:
                nosubject_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'nosubject')
                if not os.path.exists(nosubject_dir):
                    os.makedirs(nosubject_dir)
                if 'mcap' in results:
                    # Copy the multi-sheet file as Market_Cap.xlsx for backward compatibility
                    market_cap_dest = os.path.join(nosubject_dir, 'Market_Cap.xlsx')
                    shutil.copy2(excel_path, market_cap_dest)
                    add_log(f"✓ Multi-sheet Excel copied to {market_cap_dest}")
            except Exception as exc:
                add_log(f"⚠️ Could not copy Excel to nosubject/: {exc}")

            # Send the single Excel file (no ZIP needed!)
            final_memory = get_memory_usage_mb()
            total_elapsed = time.perf_counter() - stage_start
            add_log(f"Export completed: {total_elapsed:.2f}s, Peak memory: {final_memory:.1f}MB, File size: {excel_size_mb:.1f}MB")
            
            response = send_file(
                excel_path,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=excel_filename
            )
            response.headers['Content-Disposition'] = f"attachment; filename={excel_filename}"
            
            if logs:
                safe_log = ' | '.join(logs)
                safe_log = safe_log.encode('ascii', errors='ignore').decode('ascii')
                response.headers['X-Export-Log'] = safe_log

            print(f"[consolidate-saved][done] id={req_id} sheets={list(results.keys())} elapsed={total_elapsed:.2f}s memory={final_memory:.1f}MB size={excel_size_mb:.1f}MB")

            response.call_on_close(lambda: [shutil.rmtree(work_dir, ignore_errors=True), gc.collect()])
            return response

        except Exception as e:
            if work_dir:
                shutil.rmtree(work_dir, ignore_errors=True)
            error_msg = f"id={req_id} {e}" if req_id else str(e)
            print(f"[consolidate-saved][error] {error_msg}")
            return jsonify({'error': str(e)}), 500

    except Exception as e:
        if work_dir:
            shutil.rmtree(work_dir, ignore_errors=True)
        print(f"[consolidate-saved][error] {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/nifty-indices/fetch-and-store', methods=['POST'])
def fetch_and_store_nifty_indices():
    """
    Fetch Nifty indices from CSV files and store in MongoDB.
    This should be called when user clicks the 'Refresh Indices' button.
    """
    if nifty_indices_collection is None:
        return jsonify({'error': 'Database not connected'}), 500
    
    try:
        print("[fetch-and-store-indices] Starting index fetch from live NSE API...")
        
        # We'll fetch the same fundamental indices plus any extras the user might want
        target_indices = [
            'NIFTY 50', 'NIFTY NEXT 50', 'NIFTY MIDCAP 150', 
            'NIFTY SMALLCAP 250', 'NIFTY MICROCAP 250', 'NIFTY 500'
        ]
        
        # Use our new mapping builder that grabs live FF/MC data
        index_mapping, live_data_mapping, errors = build_symbol_index_map(target_indices)
        
        if not index_mapping:
            return jsonify({'error': 'Failed to fetch any indices from NSE API', 'details': errors}), 500
        
        # --- MANUAL OVERRIDES PHASE ---
        print(f"[fetch-and-store-indices] Applying manual overrides...")
        # 1. ADD: Akzo Nobel (AKZOINDIA) to Smallcap 250 if missing
        if 'AKZOINDIA' not in index_mapping:
            index_mapping['AKZOINDIA'] = ['NIFTY SMALLCAP 250']
        elif 'NIFTY SMALLCAP 250' not in index_mapping['AKZOINDIA']:
            index_mapping['AKZOINDIA'].append('NIFTY SMALLCAP 250')
            
        # 2. REMOVE: Extras from Smallcap 250 as requested by user
        extras_to_remove = {'BASF', 'RELINFRA', 'SUNDRMFAST', 'VENTIVE', 'ASTRAZEN'}
        for sym_rem in extras_to_remove:
            sym_rem_up = sym_rem.upper()
            if sym_rem_up in index_mapping:
                # Remove index from list
                if 'NIFTY SMALLCAP 250' in index_mapping[sym_rem_up]:
                    index_mapping[sym_rem_up].remove('NIFTY SMALLCAP 250')
                # If no indices left for this symbol, remove it from mapping entirely
                if not index_mapping[sym_rem_up]:
                    del index_mapping[sym_rem_up]
        
        # 3. Add Microcap hardcoded symbols (moved earlier for unified bulk ops)
        for _sym in ['GANECOS', 'ALLCARGO']:
            _sym_up = _sym.upper()
            if _sym_up not in index_mapping:
                index_mapping[_sym_up] = ['NIFTY MICROCAP 250']
            elif 'NIFTY MICROCAP 250' not in index_mapping[_sym_up]:
                index_mapping[_sym_up].append('NIFTY MICROCAP 250')
        # ------------------------------
        
        # Prepare bulk operations for MongoDB
        bulk_operations = []
        timestamp = datetime.now()
        
        for symbol, indices in index_mapping.items():
            # Standardize symbol for storage
            symbol = str(symbol).strip().upper()
            if symbol == 'PERMITTED':
                continue
            # Filter out "Permitted" index
            filtered_indices = [idx for idx in indices if str(idx).strip().upper() != 'PERMITTED']
            if not filtered_indices:
                continue
                
            # Get live data if available
            live_vals = live_data_mapping.get(symbol, {})
            live_mc = live_vals.get('mc')
            live_ff = live_vals.get('ff')
            
            bulk_operations.append(
                UpdateOne(
                    {'symbol': symbol},
                    {
                        '$set': {
                            'symbol': symbol,
                            'indices': filtered_indices,
                            'primary_index': filtered_indices[0],
                            'live_mc': live_mc,
                            'live_ff': live_ff,
                            'last_updated': timestamp
                        }
                    },
                    upsert=True
                )
            )
        
        # Execute bulk write (with clear first to remove stale data)
        if bulk_operations:
            print(f"[fetch-and-store-indices] Clearing existing indices to prevent stale data...")
            nifty_indices_collection.delete_many({})
            
            result = nifty_indices_collection.bulk_write(bulk_operations, ordered=True)
            print(f"[fetch-and-store-indices] ✓ synchronized {len(index_mapping)} symbols with DB")
            print(f"[fetch-and-store-indices] Matched: {result.matched_count}, Modified: {result.modified_count}, Upserted: {result.upserted_count}")
            
            # Count symbols per index
            index_counts = {}
            for indices in index_mapping.values():
                for idx in indices:
                    index_counts[idx] = index_counts.get(idx, 0) + 1
            
            return jsonify({
                'success': True,
                'total_symbols': len(index_mapping),
                'timestamp': timestamp.isoformat(),
                'index_distribution': index_counts,
                'stats': {
                    'matched': result.matched_count,
                    'modified': result.modified_count,
                    'upserted': result.upserted_count
                }
            })
        else:
            return jsonify({'error': 'No data to store'}), 400
            
    except Exception as e:
        print(f"[fetch-and-store-indices] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/nifty-indices/status', methods=['GET'])
def get_nifty_indices_status():
    """
    Get the status of stored Nifty indices in DB.
    Returns count of symbols and last update timestamp.
    """
    if nifty_indices_collection is None:
        return jsonify({'error': 'Database not connected'}), 500
    
    try:
        total_count = nifty_indices_collection.count_documents({})
        
        # Get last updated timestamp
        latest_doc = nifty_indices_collection.find_one(
            {},
            sort=[('last_updated', -1)]
        )
        
        last_updated = latest_doc.get('last_updated') if latest_doc else None
        
        # Count by index
        pipeline = [
            {'$unwind': '$indices'},
            {'$group': {
                '_id': '$indices',
                'count': {'$sum': 1}
            }},
            {'$sort': {'_id': 1}}
        ]
        
        index_distribution = {}
        for doc in nifty_indices_collection.aggregate(pipeline):
            index_distribution[doc['_id']] = doc['count']
        
        return jsonify({
            'total_symbols': total_count,
            'last_updated': last_updated.isoformat() if last_updated else None,
            'index_distribution': index_distribution,
            'has_data': total_count > 0
        })
        
    except Exception as e:
        print(f"[nifty-indices-status] Error: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/nifty-indices/get-symbol-indices', methods=['GET'])
def get_symbol_indices():
    """
    Get indices for specific symbols from DB.
    Query param: symbols (comma-separated)
    """
    if nifty_indices_collection is None:
        return jsonify({'error': 'Database not connected'}), 500
    
    symbols_param = request.args.get('symbols', '')
    if not symbols_param:
        return jsonify({'error': 'Missing symbols parameter'}), 400
    
    try:
        symbols = [s.strip() for s in symbols_param.split(',') if s.strip()]
        
        results = {}
        for doc in nifty_indices_collection.find({'symbol': {'$in': symbols}}):
            results[doc['symbol']] = {
                'indices': doc.get('indices', []),
                'primary_index': doc.get('primary_index'),
                'last_updated': doc.get('last_updated').isoformat() if doc.get('last_updated') else None
            }
        
        return jsonify(results)
        
    except Exception as e:
        print(f"[get-symbol-indices] Error: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/db-status', methods=['GET'])
def get_db_status():
    """Get database collection counts and status."""
    if db is None:
        return jsonify({'error': 'Database not connected'}), 500
    try:
        status = {}
        for coll_name in db.list_collection_names():
            status[coll_name] = db[coll_name].count_documents({})
        return jsonify({
            'status': 'connected',
            'collections': status,
            'timestamp': datetime.now().isoformat()
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/db-prune', methods=['POST'])
def prune_db():
    """Prune old database entries to free up space."""
    if db is None:
        return jsonify({'error': 'Database not connected'}), 500
    try:
        data = request.get_json() or {}
        days = int(data.get('days', 60))
        
        # 1. Clear excel_results
        res_excel = db['excel_results'].delete_many({})
        
        # 2. Clear old bhavcache
        cutoff_date = datetime.now() - timedelta(days=days)
        cutoff_iso = cutoff_date.strftime('%Y-%m-%d')
        res_cache = db['bhavcache'].delete_many({'date': {'$lt': cutoff_iso}})
        
        return jsonify({
            'message': 'Database pruned successfully',
            'deleted_excel_results': res_excel.deleted_count,
            'deleted_old_cache': res_cache.deleted_count,
            'pruned_before': cutoff_iso
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    import os
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
